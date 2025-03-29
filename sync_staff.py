import os
import random
import string
import json
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
from faker import Faker

from config import *
from objects.driver import Driver



def initialize_driver_timetable(
    file_path="Lenh_Dieu_Xe.xlsx", sheet_name="Driver_Timetable", is_testing=False
):
    # Kiểm tra file tồn tại hay không
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()  # Nếu không có file, tạo mới workbook

    # Nếu is_testing là True, xóa sheet nếu tồn tại
    if is_testing and sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # Kiểm tra sheet "Driver_Timetable" có tồn tại không
    if sheet_name not in wb.sheetnames:
        # Nếu không có, tạo mới sheet
        ws = wb.create_sheet(sheet_name)
        # Nếu đây là file mới, xóa sheet mặc định
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    else:
        # Nếu có, lấy sheet hiện có
        ws = wb[sheet_name]

    # Gộp ô từ A1 đến AY1 (do thêm cột STT, tăng từ AX lên AY)
    ws.merge_cells("A1:AY1")
    ws["A1"] = "Dữ liệu gộp từ A1 đến AY1"

    # Thêm tiêu đề cho A2, B2, C2 (STT, Tài xế, Số điện thoại)
    ws["A2"] = "STT"
    ws["B2"] = "Tài xế"
    ws["C2"] = "Số điện thoại"

    # Gộp các ô ở hàng 2 (dịch sang phải 1 cột)
    ws.merge_cells("D2:O2")  # Trước là C2:N2
    ws["D2"] = "Gộp 1"

    ws.merge_cells("P2:AA2")  # Trước là O2:Z2
    ws["P2"] = "Gộp 2"

    ws.merge_cells("AB2:AM2")  # Trước là AA2:AL2
    ws["AB2"] = "Gộp 3"

    ws.merge_cells("AN2:AY2")  # Trước là AM2:AX2
    ws["AN2"] = "Gộp 4"

    # Chỉnh độ rộng cột
    ws.column_dimensions["A"].width = 5  # Cột A (STT): 5 units
    ws.column_dimensions["B"].width = 20  # Cột B (Tài xế): 20 units
    ws.column_dimensions["C"].width = 15  # Cột C (Số điện thoại): 15 units
    for col in range(4, 52):  # Từ cột D đến AY (dịch sang phải 1 cột)
        if col <= 26:
            col_letter = chr(64 + col)
        else:
            col_letter = "A" + chr(64 + col - 26)
        ws.column_dimensions[col_letter].width = 2  # 2 inches

    # Định nghĩa màu
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )  # Màu vàng
    orange_fill = PatternFill(
        start_color="FFA500", end_color="FFA500", fill_type="solid"
    )  # Màu cam

    # Tô màu từ hàng 3 trở đi (dịch sang phải 1 cột)
    for row in range(3, 101):  # Từ hàng 3 đến hàng 100
        ws[f"D{row}"].fill = yellow_fill  # Cột D (trước là C)
        ws[f"P{row}"].fill = yellow_fill  # Cột P (trước là O)
        ws[f"AB{row}"].fill = yellow_fill  # Cột AB (trước là AA)
        ws[f"AN{row}"].fill = yellow_fill  # Cột AN (trước là AM)

        ws[f"O{row}"].fill = orange_fill  # Cột O (trước là N)
        ws[f"AA{row}"].fill = orange_fill  # Cột AA (trước là Z)
        ws[f"AM{row}"].fill = orange_fill  # Cột AM (trước là AL)
        ws[f"AY{row}"].fill = orange_fill  # Cột AY (trước là AX)

    # Lưu file
    wb.save(file_path)
    if is_testing:
        print(
            f"Đã chạy ở chế độ testing: xóa và tạo lại sheet '{sheet_name}' trong {file_path}"
        )
    else:
        print(f"Đã khởi tạo/cập nhật sheet '{sheet_name}' trong {file_path}")


def initialize_driver_list(
    filename="Lenh_Dieu_Xe.xlsx", sheet_name="Tai_Xe", is_testing=False
):
    # Danh sách cột theo thứ tự (đã bỏ cột "available")
    columns = [
        "stt",
        "name",
        "cccd",
        "vehicle_id",
        "phone_number",
        "vehicle_load",
    ]

    # Thiết lập độ rộng cho từng cột (đơn vị inches, bỏ "available")
    width_settings = {
        "stt": 3.5,
        "name": 20,
        "cccd": 20,
        "vehicle_id": 15,
        "phone_number": 15,
        "vehicle_load": 12,
    }

    # Kiểm tra file, nếu không tồn tại thì tạo mới workbook
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        # Xóa sheet mặc định nếu có
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # Nếu is_testing là True, xóa sheet Tai_Xe nếu tồn tại
    if is_testing and sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # Kiểm tra sheet Tai_Xe, nếu chưa có thì tạo mới
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]

    # Gộp ô A1:F1 (điều chỉnh từ G1 thành F1 vì bỏ cột "available")
    ws.merge_cells("A1:F1")
    ws["A1"] = "Thông tin tài xế"  # Có thể thay đổi tiêu đề theo ý muốn

    # Đặt header ở dòng 2
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=2, column=col_idx).value = col_name

    # Đặt độ rộng cho từng cột
    col_letters = ["A", "B", "C", "D", "E", "F"]  # Điều chỉnh từ 7 cột thành 6 cột
    for letter, field in zip(col_letters, width_settings):
        ws.column_dimensions[letter].width = width_settings[field]

    # Lưu file
    wb.save(filename)
    if is_testing:
        print(
            f"Đã chạy ở chế độ testing: xóa và tạo lại sheet '{sheet_name}' trong file '{filename}'."
        )
    else:
        print(
            f"Đã tạo/cập nhật sheet '{sheet_name}' trong file '{filename}' với header ở dòng 2."
        )

def sample_drivers(filename="Lenh_Dieu_Xe.xlsx", sheet_name="Tai_Xe", number_of_drivers=5):
    # Khởi tạo Faker với locale tiếng Việt và seed cố định để kết quả không đổi
    fake = Faker("vi_VN")
    Faker.seed(42)  # Seed cố định để đảm bảo kết quả nhất quán

    # Kiểm tra file tồn tại
    if not os.path.exists(filename):
        print(
            f"File '{filename}' không tồn tại. Vui lòng chạy initialize_driver_schedule trước."
        )
        return

    # Đọc danh sách tải trọng xe từ file vehicle.json
    vehicle_file = "data/vehicle.json"
    if not os.path.exists(vehicle_file):
        print(f"File '{vehicle_file}' không tồn tại. Tạo danh sách mặc định.")
        vehicle_loads = [1.5, 2.0, 3.5, 5.0, 10.0]  # Mặc định nếu file không tồn tại
    else:
        with open(vehicle_file, "r", encoding="utf-8") as f:
            vehicle_loads = json.load(f)
        print(f"Đã đọc danh sách tải trọng xe từ {vehicle_file}: {vehicle_loads}")

    # Mở workbook
    wb = openpyxl.load_workbook(filename)

    # Kiểm tra sheet tồn tại
    if sheet_name not in wb.sheetnames:
        print(
            f"Sheet '{sheet_name}' không tồn tại. Vui lòng chạy initialize_driver_schedule trước."
        )
        return

    ws = wb[sheet_name]

    # Xác định dòng bắt đầu (sau dòng 2 là header)
    start_row = 3

    # Nếu đã có dữ liệu, tìm dòng trống tiếp theo
    if ws.max_row >= start_row:
        start_row = ws.max_row + 1

    # Tạo danh sách tài xế
    for i in range(number_of_drivers):
        row = start_row + i
        # STT
        cell_a = ws[f"A{row}"]
        cell_a.value = str(i + 1)  # Chuyển thành chuỗi để đảm bảo kiểu text
        cell_a.number_format = '@'  # Định dạng text

        # Tên tài xế (sử dụng Faker)
        cell_b = ws[f"B{row}"]
        cell_b.value = fake.name()
        cell_b.number_format = '@'  # Định dạng text

        # CCCD (12 chữ số ngẫu nhiên từ Faker)
        cell_c = ws[f"C{row}"]
        cell_c.value = str(fake.unique.random_number(digits=12, fix_len=True))
        cell_c.number_format = '@'  # Định dạng text

        # Vehicle ID (biển số xe, ví dụ: 29A-12345)
        cell_d = ws[f"D{row}"]
        cell_d.value = f"29A-{fake.unique.random_int(min=10000, max=99999):05d}"
        cell_d.number_format = '@'  # Định dạng text

        # Số điện thoại (10 chữ số, bắt đầu bằng 0)
        cell_e = ws[f"E{row}"]
        cell_e.value = fake.phone_number()
        cell_e.number_format = '@'  # Định dạng text

        # Tải trọng xe (lấy từ danh sách vehicle_loads)
        vehicle_load = vehicle_loads[i % len(vehicle_loads)] if vehicle_loads else 1.5
        cell_f = ws[f"F{row}"]
        cell_f.value = str(vehicle_load)  # Chuyển thành chuỗi để đảm bảo kiểu text
        cell_f.number_format = '@'  # Định dạng text

    # Định dạng lại header (dòng 2) thành text nếu cần
    headers = ["A2", "B2", "C2", "D2", "E2", "F2"]
    for cell_ref in headers:
        cell = ws[cell_ref]
        if cell.value:  # Nếu ô header đã có giá trị
            cell.number_format = '@'

    # Lưu file
    wb.save(filename)
    print(
        f"Đã thêm {number_of_drivers} tài xế vào sheet '{sheet_name}' trong file '{filename}' bắt đầu từ dòng {start_row}. Tất cả ô được định dạng kiểu Text."
    )


def copy_driver_data_to_timetable(
    file_path="Lenh_Dieu_Xe.xlsx",
    source_sheet="Tai_Xe",
    target_sheet="Driver_Timetable",
):
    # Kiểm tra file tồn tại
    if not os.path.exists(file_path):
        print(f"File '{file_path}' không tồn tại. Vui lòng tạo file trước.")
        return

    # Mở workbook
    wb = openpyxl.load_workbook(file_path)

    # Kiểm tra sự tồn tại của cả hai sheet
    if source_sheet not in wb.sheetnames:
        print(f"Sheet '{source_sheet}' không tồn tại trong file '{file_path}'.")
        return
    if target_sheet not in wb.sheetnames:
        print(f"Sheet '{target_sheet}' không tồn tại trong file '{file_path}'.")
        return

    # Lấy hai sheet
    ws_source = wb[source_sheet]  # Sheet Tai_Xe
    ws_target = wb[target_sheet]  # Sheet Driver_Timetable

    # Xác định các cột nguồn và đích
    source_columns = {"stt": "A", "name": "B", "phone_number": "E"}  # Từ Tai_Xe
    target_columns = {
        "STT": "A",
        "Tài xế": "B",
        "Số điện thoại": "C",
    }  # Sang Driver_Timetable

    # Bắt đầu từ dòng 3 (dòng dữ liệu đầu tiên)
    start_row = 3

    # Sao chép dữ liệu
    for row in range(start_row, ws_source.max_row + 1):
        # Lấy giá trị từ Tai_Xe
        stt = ws_source[f"{source_columns['stt']}{row}"].value
        name = ws_source[f"{source_columns['name']}{row}"].value
        phone = ws_source[f"{source_columns['phone_number']}{row}"].value

        # Ghi vào Driver_Timetable
        ws_target[f"{target_columns['STT']}{row}"].value = stt
        ws_target[f"{target_columns['Tài xế']}{row}"].value = name
        ws_target[f"{target_columns['Số điện thoại']}{row}"].value = phone

    # Lưu file
    wb.save(file_path)
    print(
        f"Đã sao chép dữ liệu STT, Tài xế, Số điện thoại từ sheet '{source_sheet}' sang sheet '{target_sheet}' trong file '{file_path}'."
    )


def check_driver_availability(
    file_path="Lenh_Dieu_Xe.xlsx", sheet_name="Driver_Timetable"
):
    # Kiểm tra file tồn tại
    if not os.path.exists(file_path):
        print(f"File '{file_path}' không tồn tại.")
        return None

    # Mở workbook
    wb = openpyxl.load_workbook(file_path)

    # Kiểm tra sheet tồn tại
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' không tồn tại trong file '{file_path}'.")
        return None

    ws = wb[sheet_name]

    # Danh sách cột từ D đến AY (48 cột, mỗi cột là 30 phút)
    col_letters = []
    for col in range(4, 52):  # Từ cột D (4) đến AY (51)
        if col <= 26:
            col_letter = chr(64 + col)
        else:
            col_letter = "A" + chr(64 + col - 26)
        col_letters.append(col_letter)

    # Màu rảnh: không tô màu (None), vàng (FFFFFF00), cam (FFFFA500)
    free_colors = [None, "FFFF00", "FFA500"]

    # Danh sách kết quả
    driver_availability = {}

    # Duyệt qua các dòng từ 3 trở đi (dòng dữ liệu tài xế)
    for row in range(3, ws.max_row + 1):
        driver_name = ws[f"B{row}"].value  # Cột B: Tên tài xế
        phone_number = ws[f"C{row}"].value  # Cột C: Số điện thoại
        if not driver_name or not phone_number:  # Nếu thiếu tên hoặc số điện thoại, bỏ qua
            continue

        # Kết hợp tên và số điện thoại làm key
        driver_key = f"{driver_name} - {phone_number}"

        # Danh sách trạng thái rảnh cho từng khung giờ
        free_slots = []
        for col_idx, col_letter in enumerate(col_letters):
            cell = ws[f"{col_letter}{row}"]
            fill_color = (
                cell.fill.start_color.index
                if cell.fill.patternType == "solid"
                else None
            )
            # print(fill_color if fill_color is None else fill_color[2:],type(fill_color))
            fill_color = fill_color if fill_color is None else fill_color[2:]
            if fill_color not in free_colors:
                print(f"Cell {col_letter}{row} is not free with color {fill_color}")
            is_free = fill_color in free_colors
            free_slots.append(is_free)

        # Tính các khoảng thời gian rảnh
        availability = []
        start_time = None
        for i in range(len(free_slots)):
            if free_slots[i] and start_time is None:  # Bắt đầu khoảng rảnh
                start_time = i * 0.5  # Chuyển sang giờ (0.5 = 30 phút)
            elif not free_slots[i] and start_time is not None:  # Kết thúc khoảng rảnh
                end_time = i * 0.5
                availability.append((start_time, end_time))
                start_time = None
            # Trường hợp cuối ngày
            if i == len(free_slots) - 1 and start_time is not None:
                availability.append((start_time, 24.0))

        # Thêm vào kết quả với key là tên + số điện thoại
        driver_availability[driver_key] = availability

    # Đóng workbook
    wb.close()

    # In kết quả
    print(f"Thời gian rảnh của các tài xế trong sheet '{sheet_name}':")
    for driver, times in driver_availability.items():
        print(f"- {driver}: {times}")
    print(driver_availability)
    return driver_availability

def driver_excel_2_csv(excel_file = "data/input/Lenh_Dieu_Xe.xlsx",sheet_name = "Tai_Xe",json_file = "data/drivers.json",is_check_driver_availability=False, checkday = TODAY):    
    wb = openpyxl.load_workbook(excel_file)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' không tồn tại trong file '{excel_file}'.")
        return
    
    ws = wb[sheet_name]
    
    drivers = []
    for row in range(3, ws.max_row + 1):
        name = ws[f"B{row}"].value
        cccd = ws[f"C{row}"].value
        vehicle_id = ws[f"D{row}"].value
        phone_number = ws[f"E{row}"].value
        vehicle_load = ws[f"F{row}"].value
        
        if name and cccd and vehicle_id and phone_number and vehicle_load:
            driver = Driver(
                name=str(name),
                cccd=str(cccd),
                vehicle_id=str(vehicle_id),
                phone_number=str(phone_number),
                vehicle_load=float(vehicle_load)
            )
            drivers.append(driver)
    
    if is_check_driver_availability:
        availability = check_driver_availability(file_path=excel_file,sheet_name=f"Driver_Timetable_{checkday}")
        if availability:
            for driver in drivers:
                driver_key = f"{driver.name} - {driver.phone_number}"
                if driver_key in availability:
                    driver.update_available_times(checkday, availability[driver_key])
    print(drivers[0].available_times, "()"*100)
    
    if os.path.exists(json_file):
        with open(json_file, "r", encoding="utf-8") as f:
            existing_drivers = json.load(f)
            existing_cccds = {driver["cccd"] for driver in existing_drivers}
        
        updated_drivers = existing_drivers[:]
        for driver in drivers:
            if driver.cccd not in existing_cccds:
                print(f"Tài xế mới: {driver.name} (CCCD: {driver.cccd}) được thêm vào danh sách.")
                updated_drivers.append(driver.to_dict())
            #cập nhật update_available_times cho updated_drivers
        # ✅ BỔ SUNG: cập nhật available_times
        if is_check_driver_availability and availability:
            for i,driver_dict in enumerate(updated_drivers):
                driver_key = f"{driver_dict['name']} - {driver_dict['phone_number']}"
                if driver_key in availability:
                    driver = Driver.from_dict(driver_dict)
                    driver.update_available_times(checkday, availability[driver_key])
                    driver_dict =driver.to_dict()
                    updated_drivers[i] = driver_dict
        
        with open(json_file, "w", encoding="utf-8") as f:
            json.dump(updated_drivers, f, ensure_ascii=False, indent=4)
        
        for row_idx, driver_dict in enumerate(updated_drivers, start=3):
            cell_b = ws[f"B{row_idx}"]
            cell_c = ws[f"C{row_idx}"]
            cell_d = ws[f"D{row_idx}"]
            cell_e = ws[f"E{row_idx}"]
            cell_f = ws[f"F{row_idx}"]
            
            cell_b.value = driver_dict["name"]
            cell_c.value = driver_dict["cccd"]
            cell_d.value = driver_dict["vehicle_id"]
            cell_e.value = driver_dict["phone_number"]
            cell_f.value = driver_dict["vehicle_load"]
            
            # Định dạng các ô có số 0 ở đầu thành chuỗi
            cell_c.number_format = "@"  # Định dạng cccd thành text
            cell_e.number_format = "@"  # Định dạng phone_number thành text
    
    else:
        seen_cccds = {}
        seen_vehicle_ids = {}
        unique_drivers = []
        
        for driver in drivers:
            if driver.cccd not in seen_cccds:
                seen_cccds[driver.cccd] = driver
            if driver.vehicle_id not in seen_vehicle_ids:
                seen_vehicle_ids[driver.vehicle_id] = driver
                unique_drivers.append(driver)
            else:
                print(f"Trùng vehicle_id {driver.vehicle_id}, chỉ giữ tài xế đầu tiên.")
        
        driver_dicts = [d.to_dict() for d in unique_drivers]
        with open(json_file, "w", encoding="utf-8") as f:
            json.dump(driver_dicts, f, ensure_ascii=False, indent=4)
        
        for row_idx, driver_dict in enumerate(driver_dicts, start=3):
            cell_b = ws[f"B{row_idx}"]
            cell_c = ws[f"C{row_idx}"]
            cell_d = ws[f"D{row_idx}"]
            cell_e = ws[f"E{row_idx}"]
            cell_f = ws[f"F{row_idx}"]
            
            cell_b.value = driver_dict["name"]
            cell_c.value = driver_dict["cccd"]
            cell_d.value = driver_dict["vehicle_id"]
            cell_e.value = driver_dict["phone_number"]
            cell_f.value = driver_dict["vehicle_load"]
            
            # Định dạng các ô có số 0 ở đầu thành chuỗi
            cell_c.number_format = "@"  # Định dạng cccd thành text
            cell_e.number_format = "@"  # Định dạng phone_number thành text
    
    wb.save(excel_file)
    wb.close()
    
    print(f"Đã xử lý dữ liệu tài xế từ '{excel_file}' và lưu vào '{json_file}'.")
    if is_check_driver_availability:
        print("Đã kiểm tra và cập nhật thời gian rảnh cho các tài xế.")

# Gọi hàm để thực thi
if __name__ == "__main__":
    # # Chạy bình thường
    # # Chạy ở chế độ testing (xóa sheet Tai_Xe nếu có và tạo lại)
    # initialize_driver_list(is_testing=True)
    # initialize_driver_timetable(is_testing=True)
    # sample_drivers(number_of_drivers=41)  # Thêm tài xế mẫu
    # copy_driver_data_to_timetable()
    # # check_driver_availability()
    # driver_excel_2_csv(is_check_driver_availability=True)
    driver_excel_2_csv(
        excel_file="data/input/Lenh_Dieu_Xe.xlsx",
        sheet_name="Tai_Xe",
        json_file="data/drivers.json",
        is_check_driver_availability=True,
        checkday=DATES[0]
    )