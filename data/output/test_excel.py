import json
import pandas as pd
import datetime
import os

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_dummy_driver_mapping(num_vehicals, script_dir):
    # Tạo mapping dummy cho vehicle từ vehicle_0 đến vehicle_{num_vehicals-1}
    data = []
    for i in range(num_vehicals):
        vehicle_id = f"vehicle_{i}"
        driver_name = f"Driver_{i}"          # Ví dụ: Driver_0, Driver_1,...
        phone_number = f"09000000{i:02d}"      # Ví dụ: 0900000000, 0900000001,...
        data.append({
            'vehicle_id': vehicle_id,
            'driver_name': driver_name,
            'phone_number': phone_number
        })
    
    mapping_df = pd.DataFrame(data)
    mapping_path = os.path.join(script_dir, 'driver_mapping.csv')
    mapping_df.to_csv(mapping_path, index=False)
    print(f"[DEBUG] Dummy driver mapping file '{mapping_path}' đã được tạo.")

def format_excel_file(file_path):
    """
    Mở file Excel bằng openpyxl và áp dụng định dạng:
      - Header: in đậm, căn giữa, nền màu, border.
      - Các ô dữ liệu: căn giữa, border, wrap_text.
      - Cài đặt chiều rộng cột (ở đây áp dụng động cho tất cả cột).
    """
    print(f"[DEBUG] Bắt đầu format file: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active

    # Định nghĩa format cho header
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="D7E4BC")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Format header (dòng 1)
    max_col = ws.max_column
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Format các ô dữ liệu (từ dòng 2 trở đi)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    # Cài đặt chiều rộng cho các cột (có thể điều chỉnh linh hoạt)
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        # Tuỳ chỉnh độ rộng, ở đây đặt rộng 25 để có chỗ cho multiline
        ws.column_dimensions[col_letter].width = 25

    wb.save(file_path)
    print(f"[DEBUG] Đã định dạng file Excel: {file_path}")


##############################################

def read_json_output_file(filename):
    # Xác định thư mục chứa file test_excel.py
    script_dir = os.path.dirname(__file__)
    
    print(f"[DEBUG] Bắt đầu đọc file JSON: {filename}")
    # Bước 1: Đọc file JSON (file .txt nhưng nội dung JSON hợp lệ)
    with open(filename, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Đọc file mapping cho tài xế
    mapping_path = os.path.join(script_dir, 'driver_mapping.csv')
    print(f"[DEBUG] Đọc driver mapping từ: {mapping_path}")
    mapping_df = pd.read_csv(mapping_path)

    # Đọc file mapping cho địa điểm
    destinations_path = os.path.abspath(os.path.join(script_dir, '..', 'destinations.csv'))
    print(f"[DEBUG] Đọc destinations từ: {destinations_path}")
    destinations_df = pd.read_csv(destinations_path)

    # Tạo thư mục xuất file Excel: Project-70m/data/output/output_excel
    output_dir = os.path.join(script_dir, 'output_excel')
    os.makedirs(output_dir, exist_ok=True)
    print(f"[DEBUG] Thư mục output_excel: {output_dir}")
    
    # Giả định data[0] chứa meta và các phần còn lại là dữ liệu theo ngày
    meta_info = data[0]
    print(f"[DEBUG] Meta info: {meta_info}")

    for day_data in data[1:]:
        date = day_data['date']
        vehicles = day_data['vehicles']
        print(f"\n[DEBUG] Xử lý ngày: {date}, số xe: {len(vehicles)}")

        # Thu thập dữ liệu cho mỗi vehicle: driver_name, phone, distance, stops
        vehicles_info = []
        max_stops = 0  # xác định số điểm dừng tối đa

        for vehicle_id, vehicle_info in vehicles.items():
            # Lấy distance_of_route, driver_name, phone
            distance = vehicle_info.get('distance_of_route', 0)
            driver_info = mapping_df[mapping_df['vehicle_id'] == vehicle_id]
            driver_name = driver_info['driver_name'].iloc[0] if not driver_info.empty else 'NoName'
            phone = driver_info['phone_number'].iloc[0] if not driver_info.empty else 'NoPhone'

            list_of_route = vehicle_info['list_of_route']
            stops = []

            for idx, route in enumerate(list_of_route):
                node = route.get('node')
                arrival_time = route.get('arrival_time')  # thời gian xuất phát dự kiến (ví dụ)
                must_deliver_time = "N/A"  # tuỳ bạn
                delivered = route.get('delivered', 0)
                capacity = route.get('capacity', 0)
                
                # Mã đơn hàng dummy
                order_code = f"DUMMY-{vehicle_id}-{idx+1}"

                # Tìm thông tin địa điểm
                try:
                    node_int = int(node) + 1
                except:
                    node_int = node
                dest_info = destinations_df[destinations_df['ID'] == node_int]
                destination_name = dest_info['Name'].iloc[0] if not dest_info.empty else 'UnknownPlace'

                multiline_text = (
                f"<<ID:{node}>> - {destination_name} - {arrival_time}\n"
                "__________________________________________________\n"
                f"{must_deliver_time}\n"
                "__________________________________________________\n"
                f"{delivered}/{capacity}\n"
                "__________________________________________________\n"
                f"<<order_code:{order_code}>> - {order_code}"
            )
                print(f"[DEBUG] Stop cell text: {multiline_text!r}")
                stops.append(multiline_text)

            # Cập nhật max_stops
            if len(stops) > max_stops:
                max_stops = len(stops)

            vehicles_info.append({
                "driver_name": driver_name,
                "phone": phone,
                "distance": distance,
                "stops": stops
            })

        # Bước tạo DataFrame:
        # Định nghĩa cột: 3 cột cố định + 2 cột cho mỗi stop (1 cột nội dung, 1 cột trống)
        # Nếu không muốn cột trống, bạn bỏ phần chèn cột trống.
        columns = ["Driver Name", "Phone", "Total Distance"]
        for i in range(max_stops):
            columns.append(f"Stop {i+1}")
            # Thêm 1 cột trống để tạo khoảng cách
            columns.append("---->")

        print(f"[DEBUG] columns = {columns}")

        # Tạo dữ liệu cho DataFrame (mỗi vehicle 1 dòng)
        df_rows = []
        for vinfo in vehicles_info:
            row_data = [
                vinfo["driver_name"],
                vinfo["phone"],
                vinfo["distance"]
            ]
            # Thêm các stop vào row_data
            for i, stop_text in enumerate(vinfo["stops"]):
                row_data.append(stop_text)  # cột stop
                row_data.append("")        # cột trống
            # Nếu còn thiếu stop so với max_stops => bổ sung cột rỗng
            missing_stops = max_stops - len(vinfo["stops"])
            for _ in range(missing_stops):
                row_data.append("")  # cột stop rỗng
                row_data.append("")  # cột trống rỗng
            df_rows.append(row_data)

        # Tạo DataFrame cuối cùng
        df_final = pd.DataFrame(df_rows, columns=columns)
        print(f"[DEBUG] DataFrame cuối cho ngày {date}:\n{df_final.head()}")

        # Xuất DataFrame ra Excel
        output_filename = os.path.join(output_dir, f"output_{date.replace('.', '-')}.xlsx")
        # Nếu file đã tồn tại, xóa đi để tránh lỗi Permission
        if os.path.exists(output_filename):
            os.remove(output_filename)

        df_final.to_excel(output_filename, index=False)
        print(f"[DEBUG] Đã xuất file Excel cho ngày {date} tại: {output_filename}")

        # Format file Excel
        format_excel_file(output_filename)




##################################################
import os
import re


def excel_to_json_single(excel_path: str, mapping_path: str):
    print(f"[DEBUG] Reading Excel file: {excel_path}")
    df = pd.read_excel(excel_path)

    mapping_df = pd.read_csv(mapping_path)
    print(f"[DEBUG] Loaded mapping: {mapping_path}")

    id_pattern    = re.compile(r"<<ID:(\d+)>>")
    order_pattern = re.compile(r"<<ORDER:([^>]+)>>")
    first_line_pattern = re.compile(r"<<ID:(\d+)>>\s*-\s*(.*?)\s*-\s*(\d+)")

    vehicles = {}
    for idx, row in df.iterrows():
        driver = row["Driver Name"]
        vehicle_row = mapping_df[mapping_df["driver_name"] == driver]
        vehicle_id = vehicle_row["vehicle_id"].iloc[0] if not vehicle_row.empty else None
        print(f"[DEBUG] Row {idx} → driver={driver}, vehicle_id={vehicle_id}")

        stops = []
        for col in df.columns:
            if not col.startswith("Stop") or pd.isna(row[col]):
                continue

            cell = str(row[col])
            lines = [line.strip() for line in cell.split("\n") if line.strip() and not re.match(r"^[_]+$", line)]

            if len(lines) < 1:
                print(f"[WARNING] Empty or malformed cell at row {idx}, col {col}")
                continue

            # Parse dòng đầu tiên: <<ID:...>> - Name - Time
            m = re.match(r"<<ID:(\d+)>>\s*-\s*(.*?)\s*-\s*(\d+)", lines[0])
            if not m:
                print(f"[WARNING] Bad first‑line format at row {idx}, col {col}: {lines[0]!r}")
                continue

            node = int(m.group(1))
            arrival_time = int(m.group(3))

            # Tìm dòng dạng x/y (delivered/capacity)
            delivered, capacity = 0, 0
            for line in lines:
                if re.match(r"^\d+/\d+$", line):
                    delivered, capacity = map(int, line.split("/"))
                    break
            else:
                print(f"[WARNING] Không tìm thấy dòng dạng 'x/y' tại row {idx}, col {col}")
                continue

            # Tìm dòng chứa order_code (theo format mới: <<order_code:...>>)
            order_code = ""
            for line in lines:
                if "<<order_code:" in line:
                    match = re.match(r"<<order_code:([^>]+)>>", line)
                    if match:
                        order_code = match.group(1)
                    break

            stops.append({
                "node": node,
                "arrival_time": arrival_time,
                "capacity": capacity,
                "delivered": delivered,
                "order_code": order_code
            })
            print(f"[DEBUG] Parsed stop → node={node}, arrival={arrival_time}, delivered={delivered}, capacity={capacity}, order={order_code}")


        vehicles[vehicle_id] = {
            "list_of_route": stops,
            "distance_of_route": int(row["Total Distance"])
        }

    basename = os.path.splitext(os.path.basename(excel_path))[0]
    date_str = basename.replace("output_", "").replace("-", ".")
    result = {"meta": "Multi-day routing output", "date": date_str, "vehicles": vehicles}

    script_dir = os.path.dirname(__file__)
    out_dir = os.path.join(script_dir, "json_after_reversed")
    os.makedirs(out_dir, exist_ok=True)

    output_json = os.path.join(out_dir, f"{basename}.json")
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"[DEBUG] Wrote JSON to {output_json}")
    return output_json



if __name__ == '__main__':
    # Xác định thư mục chứa file test_excel.py
    script_dir = os.path.dirname(__file__)
    # Tính đường dẫn tới file config.json (nằm 2 cấp thư mục lên, tức là Project-70m/config.json)
    config_path = os.path.join(script_dir, '..', '..', 'config.json')
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
        
    num_vehicals = config.get('NUM_OF_VEHICLES')
    print("[DEBUG] num_vehicals =", num_vehicals)
    
    # Tạo file mapping dummy nếu chưa có
    create_dummy_driver_mapping(num_vehicals, script_dir)
    
    # File "2025-03-25_12-03-07.txt" nằm cùng thư mục với test_excel.py
    txt_file = os.path.join(script_dir, "2025-03-25_12-03-07.txt")
    read_json_output_file(txt_file)
    excel_path = os.path.join(script_dir, "output_excel/output_01-03-2025.xlsx")
    mapping_path = os.path.join(script_dir, "driver_mapping.csv")
    excel_to_json_single(excel_path, mapping_path)
    
    
