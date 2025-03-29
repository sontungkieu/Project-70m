import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation


def create_dropdown_list(
    workbook,
    target_sheet_name,
    target_range,
    source_sheet_name,
    source_column="A",
    num_rows=0,
    allow_blank=True,
):
    """
    Thêm dropdown list vào sheet mục tiêu dựa trên dữ liệu từ sheet nguồn.

    Args:
        workbook: Đối tượng Workbook của openpyxl.
        target_sheet_name (str): Tên sheet sẽ thêm dropdown.
        target_range (str): Vùng ô (ví dụ: "A2:A12") được áp dụng dropdown.
        source_sheet_name (str): Tên sheet chứa dữ liệu cho dropdown.
        source_column (str): Cột chứa dữ liệu (mặc định "A").
        num_rows (int): Số dòng dữ liệu trong sheet nguồn.
        allow_blank (bool): Có cho phép giá trị rỗng hay không.
    """
    ws_target = workbook[target_sheet_name]

    # Xóa các data validation cũ (nếu có)
    if ws_target.data_validations is not None:
        ws_target.data_validations.dataValidation = []

    # Tạo data validation kiểu list với công thức tham chiếu đến sheet nguồn
    dv = DataValidation(
        type="list",
        formula1=f"'{source_sheet_name}'!${source_column}$3:${source_column}${num_rows}",
        allow_blank=allow_blank,
    )
    dv.add(target_range)
    ws_target.add_data_validation(dv)


def sync_csv_to_excel(
    csv_file,
    excel_file,
    sheet2_name="Dia_Chi",
    add_drop_down=False,
    sheet1_name="3.3",
    sheet1_range="A2:A12",
    is_get_from_csv=True,  # Thêm param mới để kiểm soát lấy dữ liệu từ CSV
    is_overwrite=True      # Điều chỉnh để kiểm soát ghi đè dropdown
):
    # Load file Excel hiện có
    wb = load_workbook(excel_file)

    # Truy cập Sheet2 (tạo mới nếu chưa tồn tại)
    if sheet2_name not in wb.sheetnames:
        ws2 = wb.create_sheet(sheet2_name)
    else:
        ws2 = wb[sheet2_name]

    # Nếu is_get_from_csv=True, lấy dữ liệu từ CSV và đồng bộ vào Sheet2
    if is_get_from_csv:
        # Đọc file CSV
        df = pd.read_csv(csv_file)

        # Xóa dữ liệu cũ trong Sheet2 (nếu có)
        ws2.delete_rows(1, ws2.max_row)

        # Điền dữ liệu từ CSV vào Sheet2 (cột A: Name, cột B: Address, cột C: Name + Address)
        for row, (name, address) in enumerate(zip(df["Name"], df["Address"]), start=1):
            ws2[f"A{row}"] = name
            ws2[f"B{row}"] = address
            ws2[f"C{row}"] = f"{name} - {address}"  # Tạo cột C bằng cách nối A và B
        print(f"Đã đồng bộ dữ liệu từ {csv_file} vào sheet '{sheet2_name}' trong {excel_file}.")
    else:
        print(f"Bỏ qua lấy dữ liệu từ CSV vì is_get_from_csv=False.")

    # Xử lý dropdown trong Sheet1
    num_rows = 1000
    if add_drop_down:
        # Kiểm tra nếu sheet1 đã có dropdown và is_overwrite=False
        ws1 = wb[sheet1_name]
        if ws1.data_validations.count > 0 and not is_overwrite:
            print(f"Dropdown trong '{sheet1_name}' đã tồn tại và is_overwrite=False, bỏ qua cập nhật dropdown.")
        else:
            create_dropdown_list(
                workbook=wb,
                target_sheet_name=sheet1_name,  # Sheet "3.3" chứa dropdown
                target_range=sheet1_range,      # Vùng ô áp dụng dropdown
                source_sheet_name=sheet2_name,  # Sheet chứa danh sách giá trị
                source_column="C",              # Sử dụng cột C làm nguồn
                num_rows=num_rows,
                allow_blank=True,
            )
            print(f"Đã cập nhật dropdown trong '{sheet1_name}' từ cột C của '{sheet2_name}'.")

    # Lưu file Excel
    wb.save(excel_file)
    print(f"Đã lưu file {excel_file} thành công!")

def excel_sheet2_to_csv(excel_file, csv_file, sheet2_name="Dia_Chi"):
    # Load file Excel
    wb = load_workbook(excel_file, read_only=True)

    # Truy cập Sheet2
    if sheet2_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet2_name}' không tồn tại trong file {excel_file}")
    ws2 = wb[sheet2_name]

    # Lấy dữ liệu từ Sheet2
    data = []
    for row in ws2.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] is not None:  # Chỉ lấy hàng có dữ liệu ở cột A
            data.append({"Name": row[0], "Address": row[1] if row[1] else ""})

    if not data:
        raise ValueError(f"Sheet2 trong {excel_file} không có dữ liệu!")

    # Tạo DataFrame và thêm cột ID
    df = pd.DataFrame(data)
    df.insert(0, "ID", range(1, len(df) + 1))  # Thêm cột ID từ 1

    # Lưu vào file CSV
    df.to_csv(csv_file, index=False, encoding="utf-8")
    print(f"Đã chuyển dữ liệu từ {excel_file}, Sheet2 sang {csv_file}!")


if __name__ == "__main__":
    # Sử dụng hàm
    csv_file = "data/destinations.csv"
    excel_file = "Lenh_Dieu_Xe.xlsx"
    excel_sheet2_to_csv(excel_file, csv_file)
    # sync_csv_to_excel(
    #     csv_file,
    #     excel_file,
    #     "Dia_Chi",
    #     add_drop_down=True,
    #     sheet1_name="3.4",
    #     sheet1_range="B5:B55",
    # )
