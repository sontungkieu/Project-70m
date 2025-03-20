import random
from copy import copy
from datetime import datetime, timedelta

import openpyxl
import pandas as pd
from faker import Faker
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from config import *
from sync_destinations import sync_csv_to_excel
from sync_staff import (
    copy_driver_data_to_timetable,
    initialize_driver_list,
    initialize_driver_timetable,
    sample_drivers,
)


def copy_excel_sheet_fully(
    file_path="Lenh_Dieu_Xe.xlsx",
    origin_sheet="Sheet1",
    sheet_name="NewSheet",
    skip_cf=False,
):
    try:
        # Load file Excel
        wb = load_workbook(file_path)

        # Kiểm tra xem origin_sheet có tồn tại không
        if origin_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{origin_sheet}' không tồn tại trong file Excel")

        # Lấy sheet gốc
        source_sheet = wb[origin_sheet]

        # Tạo sheet mới
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])  # Xóa sheet cũ nếu đã tồn tại
        target_sheet = wb.create_sheet(sheet_name)

        # Copy tất cả các ô giữ nguyên format
        for row in source_sheet.rows:
            for cell in row:
                new_cell = target_sheet[cell.coordinate]
                new_cell.value = cell.value

                # Copy định dạng
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

                # Copy comment/note
                if cell.comment:
                    new_cell.comment = copy(cell.comment)

        # Copy chiều rộng cột (đã sửa lỗi cú pháp)
        for col in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[
                col
            ].width
            target_sheet.column_dimensions[col].hidden = source_sheet.column_dimensions[
                col
            ].hidden

        # Copy chiều cao hàng và trạng thái ẩn
        for row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[
                row
            ].height
            target_sheet.row_dimensions[row].hidden = source_sheet.row_dimensions[
                row
            ].hidden

        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

        # Copy filter
        if source_sheet.auto_filter.ref:
            target_sheet.auto_filter.ref = source_sheet.auto_filter.ref

        # Copy freeze panes
        if source_sheet.freeze_panes:
            target_sheet.freeze_panes = source_sheet.freeze_panes

        # Copy conditional formatting (nếu không skip)
        if not skip_cf:
            from openpyxl.formatting.rule import Rule
            from openpyxl.styles.differential import DifferentialStyle

            for cf in source_sheet.conditional_formatting:
                for cell_range in cf.cells:
                    range_string = str(cell_range)
                    for rule in cf.rules:
                        if hasattr(rule, "dxf") and rule.dxf:
                            dxf = DifferentialStyle(
                                font=copy(rule.dxf.font) if rule.dxf.font else None,
                                border=copy(rule.dxf.border)
                                if rule.dxf.border
                                else None,
                                fill=copy(rule.dxf.fill) if rule.dxf.fill else None,
                                alignment=copy(rule.dxf.alignment)
                                if rule.dxf.alignment
                                else None,
                            )
                        else:
                            dxf = None

                        new_rule = Rule(
                            type=rule.type,
                            dxf=dxf,
                            formula=rule.formula if hasattr(rule, "formula") else None,
                            stopIfTrue=rule.stopIfTrue
                            if hasattr(rule, "stopIfTrue")
                            else None,
                            priority=rule.priority
                            if hasattr(rule, "priority")
                            else None,
                            operator=rule.operator
                            if hasattr(rule, "operator")
                            else None,
                            text=rule.text if hasattr(rule, "text") else None,
                        )
                        target_sheet.conditional_formatting.add(range_string, new_rule)

        # Copy data validation (bao gồm dropdown list) và cô lập tham chiếu
        for dv in source_sheet.data_validations.dataValidation:
            new_dv = DataValidation(
                type=dv.type,
                formula1=dv.formula1,
                formula2=dv.formula2 if dv.formula2 else None,
                allow_blank=dv.allow_blank if hasattr(dv, "allow_blank") else True,
                operator=dv.operator if hasattr(dv, "operator") else None,
            )
            new_dv.ranges = dv.ranges
            if dv.formula1 and isinstance(dv.formula1, str) and "!" in dv.formula1:
                ref_sheet, ref_range = dv.formula1.split("!")
                if ref_sheet.startswith("="):
                    ref_sheet = ref_sheet[1:]
                if ref_sheet in wb.sheetnames:
                    ref_values = []
                    for row in wb[ref_sheet][ref_range]:
                        for cell in row:
                            if cell.value:
                                ref_values.append(str(cell.value))
                    new_dv.formula1 = f'"{",".join(ref_values)}"'
            target_sheet.add_data_validation(new_dv)

        # Copy sheet properties
        target_sheet.sheet_properties.tabColor = source_sheet.sheet_properties.tabColor
        target_sheet.views = source_sheet.views

        # Lưu file
        wb.save(file_path)
        print(
            f"Đã copy sheet '{origin_sheet}' sang sheet mới '{sheet_name}' với mọi thuộc tính!"
        )

    except FileNotFoundError:
        print(f"Không tìm thấy file '{file_path}'")
    except Exception as e:
        print(f"Đã xảy ra lỗi: {str(e)}")


def copy_excel_sheet_with_format_and_filter(
    file_path="Lenh_Dieu_Xe.xlsx", origin_sheet="Sheet1", sheet_name="NewSheet"
):
    try:
        # Load file Excel
        wb = load_workbook(file_path)

        # Kiểm tra xem origin_sheet có tồn tại không
        if origin_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{origin_sheet}' không tồn tại trong file Excel")

        # Lấy sheet gốc
        source_sheet = wb[origin_sheet]

        # Tạo sheet mới
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])  # Xóa sheet cũ nếu đã tồn tại
        target_sheet = wb.create_sheet(sheet_name)

        # Copy tất cả các ô giữ nguyên format
        for row in source_sheet.rows:
            for cell in row:
                new_cell = target_sheet[cell.coordinate]
                new_cell.value = cell.value

                # Copy định dạng
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copy chiều rộng cột
        for col in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[
                col
            ].width

        # Copy chiều cao hàng
        for row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[
                row
            ].height

        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

        # Copy filter (bộ lọc)
        if source_sheet.auto_filter.ref:
            target_sheet.auto_filter.ref = source_sheet.auto_filter.ref

        # Lưu file
        wb.save(file_path)
        print(
            f"Đã copy sheet '{origin_sheet}' sang sheet mới '{sheet_name}' với định dạng và filter nguyên vẹn!"
        )

    except FileNotFoundError:
        print(f"Không tìm thấy file '{file_path}'")
    except Exception as e:
        print(f"Đã xảy ra lỗi: {str(e)}")


def copy_excel_sheet_between_files(
    source_file="data/Lenh_Dieu_Xe.xlsx",
    source_sheet="Template",
    target_file="Lenh_Dieu_Xe.xlsx",
    target_sheet_name=TODAY,
):
    try:
        # Load file nguồn
        source_wb = load_workbook(source_file)

        # Kiểm tra xem source_sheet có tồn tại không
        if source_sheet not in source_wb.sheetnames:
            raise ValueError(
                f"Sheet '{source_sheet}' không tồn tại trong file '{source_file}'"
            )

        # Lấy sheet gốc từ file nguồn
        source_sheet = source_wb[source_sheet]
        # Kiểm tra file đích, nếu không tồn tại thì tạo mới
        try:
            target_wb = load_workbook(target_file)
        except FileNotFoundError:
            # Tạo workbook mới nếu file đích không tồn tại
            target_wb = Workbook()
            # Xóa sheet mặc định được tạo bởi Workbook()
            if "Sheet" in target_wb.sheetnames:
                target_wb.remove(target_wb["Sheet"])

        # Tạo hoặc thay thế sheet mới trong file đích
        if target_sheet_name in target_wb.sheetnames:
            target_wb.remove(
                target_wb[target_sheet_name]
            )  # Xóa sheet cũ nếu đã tồn tại
        target_sheet = target_wb.create_sheet(target_sheet_name)

        # Copy tất cả các ô giữ nguyên format
        for row in source_sheet.rows:
            for cell in row:
                new_cell = target_sheet[cell.coordinate]
                new_cell.value = cell.value

                # Copy định dạng
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copy chiều rộng cột
        for col in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[
                col
            ].width

        # Copy chiều cao hàng
        for row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[
                row
            ].height

        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

        # Copy filter (bộ lọc)
        if source_sheet.auto_filter.ref:
            target_sheet.auto_filter.ref = source_sheet.auto_filter.ref

        # Lưu file đích
        target_wb.save(target_file)
        print(
            f"Đã copy sheet '{source_sheet}' từ '{source_file}' sang sheet '{target_sheet_name}' trong '{target_file}' "
            "với định dạng và filter nguyên vẹn!"
        )

    except FileNotFoundError:
        print(f"Không tìm thấy file '{source_file}'")
    except Exception as e:
        print(f"Đã xảy ra lỗi: {str(e)}")


def initialize_excel_utilities(is_recreate=False):
    # Tạo danh sách thời gian
    time_list = []
    start_time = datetime(2023, 1, 1, 0, 0)

    for i in range(24 * TIME_SCALE + 1):
        current_time = start_time + timedelta(minutes=60 / TIME_SCALE * i)
        time_str = current_time.strftime("%H:%M")
        time_list.append(time_str)
    time_list[-1] = "24:00"

    # Tạo danh sách full_range_time
    full_range_time = [
        "S(08:00->12:00)",
        "C(13:30->17:30)",
        "T(19:00->23:00)",
        "D(00:30->04:30)",
        "Báo sau",
    ]

    # Danh sách các loại xe (mới thêm)
    vehicle_types = [
        "1.4t = 9.7m3",
        "3.5t = 24.2m3",
        "5t = 26.7m3",
        "7t = 32m3",
        "9t = 38.2m3",
        "12t = 54m3",
    ]

    # Tên file
    filename = "Lenh_Dieu_Xe.xlsx"

    # Kiểm tra và tạo file nếu chưa tồn tại
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Xử lý sheet "CONFIG"
    if "CONFIG" in wb.sheetnames and is_recreate:
        # Xóa sheet cũ nếu is_recreate = True
        wb.remove(wb["CONFIG"])
        ws = wb.create_sheet("CONFIG")
    elif "CONFIG" not in wb.sheetnames:
        # Tạo sheet mới nếu chưa có
        ws = wb.create_sheet("CONFIG")
    else:
        # Sử dụng sheet hiện có nếu không recreate
        ws = wb["CONFIG"]

    # Ghi full_range_time từ ô AA1, mỗi phần tử 1 dòng
    for idx, time_range in enumerate(full_range_time, start=1):
        cell = ws[f"AA{idx}"]
        cell.value = time_range
        cell.alignment = Alignment(horizontal="left")  # Căn trái cho đẹp

    # Ghi checkbox trống ở AB1
    ws["AB1"] = "☐"  # Ký hiệu ô trống

    # Ghi checkbox đã tích ở AB2
    ws["AB2"] = "☑"  # Ký hiệu ô đã tích

    # Ghi danh sách các loại xe từ F1 trở xuống (mới thêm)
    for idx, vehicle in enumerate(vehicle_types, start=1):
        cell = ws[f"F{idx}"]
        cell.value = vehicle
        cell.alignment = Alignment(horizontal="left")  # Căn trái cho đẹp

    # Lưu file
    wb.save(filename)
    print(f"Đã tạo/thêm dữ liệu vào file {filename}, sheet CONFIG")
    print(f"Tổng số khoảng thời gian: {len(full_range_time)}")
    print(f"Tổng số loại xe: {len(vehicle_types)}")


# Gọi hàm để kiểm tra
# initialize_excel_utilities()


def init_staff(file_path="Lenh_Dieu_Xe.xlsx", number_of_random_staff=0):
    # Khởi tạo Faker với locale tiếng Việt
    fake = Faker("vi_VN")

    # Đọc sheet CONFIG từ file Excel (nếu file tồn tại)
    try:
        # Đọc file Excel, chỉ định hàng 2 (header=1) là hàng tiêu đề
        df = pd.read_excel(file_path, sheet_name="CONFIG", header=1)
        print(f"Đã đọc file {file_path} thành công.")

        # Kiểm tra tên cột thực tế
        print("Tên cột trong DataFrame:", df.columns.tolist())

        # Chuẩn hóa tên cột (loại bỏ khoảng trắng thừa, chuẩn hóa định dạng)
        df.columns = df.columns.str.strip()

        # Đảm bảo tên cột đúng với định dạng mong đợi
        expected_columns = ["STT", "Tên nhân viên", "CCCD", "Số điện thoại"]
        if not all(col in df.columns for col in expected_columns):
            print("Tên cột không khớp, chuẩn hóa lại...")
            df.columns = expected_columns  # Ép tên cột nếu cần
        df = df.dropna(how="all")  # Xóa các hàng trống

    except Exception as e:
        # Nếu file không tồn tại hoặc sheet CONFIG không có, tạo DataFrame mới
        print(f"Lỗi khi đọc file: {e}")
        print(
            f"File {file_path} không tồn tại hoặc không có sheet CONFIG. Tạo dữ liệu mới."
        )
        df = pd.DataFrame(columns=["STT", "Tên nhân viên", "CCCD", "Số điện thoại"])

    # Load workbook để định dạng
    wb = load_workbook(file_path)
    ws = wb["CONFIG"]

    # Merge cells A1:D1
    ws.merge_cells("A1:D1")

    # Gán giá trị cho header
    ws["A1"] = "Thông tin nhân viên"
    ws["A2"] = "STT"
    ws["B2"] = "Tên nhân viên"
    ws["C2"] = "CCCD"
    ws["D2"] = "Số điện thoại"

    # Đặt độ rộng cột
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    # Xóa dữ liệu cũ (trừ header) để tránh khoảng trống
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.value = None

    # Ghi lại dữ liệu hiện có từ DataFrame (nếu có)
    if not df.empty:
        for i, row in df.iterrows():
            ws.cell(row=i + 3, column=1, value=row["STT"])
            ws.cell(row=i + 3, column=2, value=row["Tên nhân viên"])
            ws.cell(row=i + 3, column=3, value=row["CCCD"])
            ws.cell(row=i + 3, column=4, value=row["Số điện thoại"])

    # Thêm random staff nếu number_of_random_staff > 0
    if number_of_random_staff > 0:
        # Lấy STT cuối cùng nếu có dữ liệu, nếu không bắt đầu từ 1
        last_stt = len(df) if not df.empty else 0

        # Tạo danh sách nhân viên random
        random_staff = []
        for i in range(number_of_random_staff):
            stt = last_stt + i + 1
            name = fake.name()
            cccd = "".join([str(random.randint(0, 9)) for _ in range(12)])  # CCCD 12 số
            phone = fake.phone_number()
            random_staff.append([stt, name, cccd, phone])

        # Chuyển random staff thành DataFrame
        random_df = pd.DataFrame(
            random_staff, columns=["STT", "Tên nhân viên", "CCCD", "Số điện thoại"]
        )

        # Gộp dữ liệu cũ và mới
        df = pd.concat([df, random_df], ignore_index=True)

        # Ghi dữ liệu mới vào worksheet ngay sau dữ liệu hiện có
        for i, row in enumerate(random_df.values, start=3 + last_stt):
            for j, value in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=value)

        print(
            f"Đã thêm {number_of_random_staff} nhân viên ngẫu nhiên vào file {file_path}"
        )
    else:
        print(
            "Không có nhân viên ngẫu nhiên nào được thêm (number_of_random_staff = 0)"
        )

    # Xóa các hàng trống ở cuối nếu có
    max_row = ws.max_row
    for row in range(max_row, 2, -1):  # Duyệt từ dưới lên, trừ header
        if all(ws.cell(row=row, column=col).value is None for col in range(1, 5)):
            ws.delete_rows(row)
        else:
            break

    # Lưu file
    wb.save(file_path)
    print(f"Đã lưu file thành công tại {file_path}")

    return df


def create_dropdowns(
    filename = "Lenh_Dieu_Xe.xlsx",
    object_sheet = TODAY,
    range_e="CONFIG!AA1:AA4656",
    range_ijk="CONFIG!AB1:AB2",
    range_g="Dia_Chi!A1:A2",
    range_h="CONFIG!B3:B100",
    range_d="CONFIG!F1:F6",
):
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    if object_sheet not in wb.sheetnames:
        ws = wb.create_sheet(object_sheet)
    else:
        ws = wb[object_sheet]

    # Xử lý range cho cột E
    if "!" in range_e:
        sheet_name, cell_range = range_e.split("!")
        source_ws = wb[sheet_name]
    else:
        source_ws = wb["CONFIG"]
        cell_range = range_e

    start_cell, end_cell = cell_range.split(":")
    full_range_time = []
    for row in source_ws[f"{start_cell}:{end_cell}"]:
        for cell in row:
            if cell.value:
                full_range_time.append(str(cell.value).strip())

    max_length = 200
    temp_values = []
    current_length = 0
    for val in full_range_time:
        val_len = len(val) + (2 if temp_values else 0)
        if current_length + val_len <= max_length:
            temp_values.append(val)
            current_length += val_len
        else:
            break
    dropdown_values_e = ",".join(temp_values)

    dv_e = DataValidation(
        type="list", formula1=f'"{dropdown_values_e}"', allow_blank=True
    )
    dv_e.add("E5:E55")
    ws.add_data_validation(dv_e)

    # Xử lý range cho cột I, J, K
    if "!" in range_ijk:
        sheet_name, cell_range = range_ijk.split("!")
        source_ws = wb[sheet_name]
    else:
        source_ws = wb["CONFIG"]
        cell_range = range_ijk

    start_cell, end_cell = cell_range.split(":")
    checkbox_values_list = []
    for row in source_ws[f"{start_cell}:{end_cell}"]:
        for cell in row:
            if cell.value:
                checkbox_values_list.append(str(cell.value).strip())
    dropdown_values_ijk = ",".join(checkbox_values_list)

    dv_ijk = DataValidation(
        type="list", formula1=f'"{dropdown_values_ijk}"', allow_blank=True
    )
    dv_ijk.add("I5:I55")
    dv_ijk.add("J5:J55")
    dv_ijk.add("K5:K55")
    ws.add_data_validation(dv_ijk)

    # Xử lý range cho cột G (từ Dia_Chi!C1:C6)
    if "!" in range_g:
        sheet_name, cell_range = range_g.split("!")
        source_ws = wb[sheet_name]
    else:
        source_ws = wb["Dia_Chi"]
        cell_range = range_g

    start_cell, end_cell = cell_range.split(":")
    address_values = []
    for row in source_ws[f"{start_cell}:{end_cell}"]:
        for cell in row:
            if cell.value:
                address_values.append(str(cell.value).strip())
    dropdown_values_g = ",".join(address_values)

    # Tạo dropdown cho cột G (mới thêm)
    print(f"Dữ liệu tìm thấy trong {range_g}: {address_values}")
    if address_values:
        dv_g = DataValidation(
            type="list", formula1=f'"{dropdown_values_g}"', allow_blank=True
        )
        dv_g.add("G5:G55")
        ws.add_data_validation(dv_g)
    else:
        print("Không có dữ liệu hợp lệ trong range C1:C6 của sheet Dia_Chi")

    # Xử lý range cho cột H
    if "!" in range_h:
        sheet_name, cell_range = range_h.split("!")
        source_ws = wb[sheet_name]
    else:
        source_ws = wb["CONFIG"]
        cell_range = range_h

    start_cell, end_cell = cell_range.split(":")
    h_values = []
    for row in source_ws[f"{start_cell}:{end_cell}"]:
        for cell in row:
            if cell.value:
                h_values.append(str(cell.value).strip())

    print(f"Dữ liệu tìm thấy trong {range_h}: {h_values}")
    if h_values:
        dropdown_values_h = ",".join(h_values)
        dv_h = DataValidation(
            type="list", formula1=f'"{dropdown_values_h}"', allow_blank=True
        )
        dv_h.add("H5:H55")
        ws.add_data_validation(dv_h)
    else:
        print("Không có dữ liệu hợp lệ trong range B3:B100 của sheet CONFIG")

    # Xử lý range cho cột D
    if "!" in range_d:
        sheet_name, cell_range = range_d.split("!")
        source_ws = wb[sheet_name]
    else:
        source_ws = wb["CONFIG"]
        cell_range = range_d

    start_cell, end_cell = cell_range.split(":")
    d_values = []
    for row in source_ws[f"{start_cell}:{end_cell}"]:
        for cell in row:
            if cell.value:
                d_values.append(str(cell.value).strip())

    print(f"Dữ liệu tìm thấy trong {range_d}: {d_values}")
    if d_values:
        dropdown_values_d = ",".join(d_values)
        dv_d = DataValidation(
            type="list", formula1=f'"{dropdown_values_d}"', allow_blank=True
        )
        dv_d.add("D5:D55")
        ws.add_data_validation(dv_d)
    else:
        print("Không có dữ liệu hợp lệ trong range F1:F6 của sheet CONFIG")

    # Lưu file
    try:
        wb.save(filename)
        print(f"Đã thêm dropdown vào sheet {TODAY} trong file {filename}")
        print(f"Dropdown cột I,J,K: {dropdown_values_ijk} từ {range_ijk}")
        # print(f"Dropdown cột G: {len(address_values)} giá trị từ {range_g}")
        print(f"Dropdown cột H: {len(h_values)} giá trị từ {range_h}")
        print(f"Dropdown cột D: {len(d_values)} giá trị từ {range_d}")
    except Exception as e:
        print(f"Lỗi khi lưu file: {e}")


def sort_sheets_by_name():
    # Tên file
    filename = "Lenh_Dieu_Xe.xlsx"

    # Mở file
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print(f"Không tìm thấy file {filename}")
        return

    # Lấy danh sách tên sheet hiện tại
    sheet_names = wb.sheetnames
    print(f"Danh sách sheet trước khi sắp xếp: {sheet_names}")

    # Sắp xếp danh sách tên sheet theo thứ tự bảng chữ cái
    sorted_sheet_names = sorted(sheet_names)
    print(f"Danh sách sheet sau khi sắp xếp: {sorted_sheet_names}")

    # Lấy danh sách các sheet hiện tại từ wb._sheets
    sheets = wb._sheets

    # Tạo dictionary ánh xạ tên sheet với đối tượng sheet
    sheet_dict = {sheet.title: sheet for sheet in sheets}

    # Thay đổi thứ tự trong wb._sheets theo sorted_sheet_names
    wb._sheets = [sheet_dict[name] for name in sorted_sheet_names]

    # Lưu file
    try:
        wb.save(filename)
        print(f"Đã sắp xếp các sheet trong file {filename} theo tên")
    except Exception as e:
        print(f"Lỗi khi lưu file: {e}")
    finally:
        wb.close()


# Gọi hàm
# sort_sheets_by_name()

if __name__ == "__main__":
    csv_file = "data/destinations.csv"
    # copy_excel_sheet_with_format_and_filter("Lenh_Dieu_Xe.xlsx", "Template", "3.4")
    copy_excel_sheet_between_files()
    sync_csv_to_excel(
        "data/destinations.csv",
        "Lenh_Dieu_Xe.xlsx",
        "Dia_Chi",
        add_drop_down=True,
        sheet1_name=TODAY,
        sheet1_range="B5:B55",
    )
    initialize_excel_utilities(is_recreate=True)
    init_staff("Lenh_Dieu_Xe.xlsx", number_of_random_staff=5)
    # create_dropdowns(range_e="CONFIG!AA1:AA4656", range_ijk="CONFIG!AB1:AB2", range_g="Dia_Chi!C1:C6")

    create_dropdowns()
    initialize_driver_list(is_testing=True)
    initialize_driver_timetable(is_testing=True)

    sample_drivers()  # Thêm 5 tài xế mẫu
    copy_driver_data_to_timetable()
    sort_sheets_by_name()
