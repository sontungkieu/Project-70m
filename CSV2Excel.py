# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.worksheet.datavalidation import DataValidation

# def sync_csv_to_excel(csv_file, excel_file, sheet2_name="Sheet2", sheet1_range="A2:A12"):
#     # Đọc file CSV
#     df = pd.read_csv(csv_file)
    
#     # Load file Excel hiện có
#     wb = load_workbook(excel_file)
    
#     # Truy cập Sheet2
#     if sheet2_name not in wb.sheetnames:
#         ws2 = wb.create_sheet(sheet2_name)
#     else:
#         ws2 = wb[sheet2_name]
    
#     # Xóa dữ liệu cũ trong Sheet2 (nếu có)
#     ws2.delete_rows(1, ws2.max_row)
    
#     # Điền dữ liệu từ CSV vào Sheet2 (cột A: Name, cột B: Address)
#     for row, (name, address) in enumerate(zip(df['Name'], df['Address']), start=1):
#         ws2[f'A{row}'] = name
#         ws2[f'B{row}'] = address
    
#     # Truy cập Sheet1 để cập nhật drop-down list
#     ws1 = wb["Sheet1"]
    
#     # Xóa data validation cũ (nếu có)
#     ws1.data_validations = None
    
#     # Tạo drop-down list mới dựa trên số dòng trong CSV
#     num_rows = len(df)
#     dv = DataValidation(type="list", formula1=f"{sheet2_name}!$A$1:$A${num_rows}", allow_blank=True)
#     dv.add(sheet1_range)  # Áp dụng cho vùng A2:A12 (hoặc tùy chỉnh)
    
#     # Thêm data validation vào Sheet1
#     ws1.add_data_validation(dv)
    
#     # Lưu file Excel
#     wb.save(excel_file)
#     print(f"Đã đồng bộ dữ liệu từ {csv_file} vào {excel_file}, Sheet2 và cập nhật drop-down list!")

# # Sử dụng hàm
# csv_file = "data/destinations.csv"
# excel_file = "kcn_dropdown.xlsx"  # File Excel đã tạo trước đó
# sync_csv_to_excel(csv_file, excel_file)
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows

def sync_csv_to_excel(csv_file, excel_file, sheet2_name="Dia_Chi", sheet1_range="A2:A12"):
    # Đọc file CSV
    df = pd.read_csv(csv_file)
    
    # Load file Excel hiện có
    wb = load_workbook(excel_file)
    
    # Truy cập Sheet2
    if sheet2_name not in wb.sheetnames:
        ws2 = wb.create_sheet(sheet2_name)
    else:
        ws2 = wb[sheet2_name]
    
    # Xóa dữ liệu cũ trong Sheet2 (nếu có)
    ws2.delete_rows(1, ws2.max_row)
    
    # Điền dữ liệu từ CSV vào Sheet2 (cột A: Name, cột B: Address)
    for row, (name, address) in enumerate(zip(df['Name'], df['Address']), start=1):
        ws2[f'A{row}'] = name
        ws2[f'B{row}'] = address
    
    # Truy cập Sheet1 để cập nhật drop-down list
    ws1 = wb["Sheet1"]
    
    if ws1.data_validations is None:
        try:
            from openpyxl.worksheet.datavalidation import DataValidationList
            ws1.data_validations = DataValidationList()
        except ImportError as e:
            print(f"Error importing DataValidationList: {e}")
            return  # Or raise the exception if you want to halt execution
    
    # Xóa data validation cũ (nếu có)
    ws1.data_validations.dataValidation = []  # Xóa danh sách validation cũ
    
    # Tạo drop-down list mới dựa trên số dòng trong CSV
    num_rows = len(df)
    dv = DataValidation(type="list", formula1=f"{sheet2_name}!$A$1:$A${num_rows}", allow_blank=True)
    dv.add(sheet1_range)  # Áp dụng cho vùng A2:A12 (hoặc tùy chỉnh)
    
    # Thêm data validation vào Sheet1
    ws1.add_data_validation(dv)
    
    # Lưu file Excel
    wb.save(excel_file)
    print(f"Đã đồng bộ dữ liệu từ {csv_file} vào {excel_file}, Sheet2 và cập nhật drop-down list!")

# Sử dụng hàm
csv_file = "data/destinations.csv"
excel_file = "kcn_dropdown.xlsx"
sync_csv_to_excel(csv_file, excel_file)