import os
import random
import string

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def initialize_driver_timetable(
    file_path="Lenh_Dieu_Xe.xlsx", sheet_name="Driver_Timetable", is_testing=False
):
    # Kiểm tra file tồn tại hay không
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()  # Nếu không có file, tạo mới workbook

    # Nếu is_testing là True, xóa sheet nếu tồn tại
    if is_testing and sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # Kiểm tra sheet "Driver_Timetable" có tồn tại không
    if sheet_name not in wb.sheetnames:
        # Nếu không có, tạo mới sheet
        ws = wb.create_sheet(sheet_name)
        # Nếu đây là file mới, xóa sheet mặc định
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    else:
        # Nếu có, lấy sheet hiện có
        ws = wb[sheet_name]

    # Gộp ô từ A1 đến AY1 (do thêm cột STT, tăng từ AX lên AY)
    ws.merge_cells("A1:AY1")
    ws["A1"] = "Dữ liệu gộp từ A1 đến AY1"

    # Thêm tiêu đề cho A2, B2, C2 (STT, Tài xế, Số điện thoại)
    ws["A2"] = "STT"
    ws["B2"] = "Tài xế"
    ws["C2"] = "Số điện thoại"

    # Gộp các ô ở hàng 2 (dịch sang phải 1 cột)
    ws.merge_cells("D2:O2")  # Trước là C2:N2
    ws["D2"] = "Gộp 1"

    ws.merge_cells("P2:AA2")  # Trước là O2:Z2
    ws["P2"] = "Gộp 2"

    ws.merge_cells("AB2:AM2")  # Trước là AA2:AL2
    ws["AB2"] = "Gộp 3"

    ws.merge_cells("AN2:AY2")  # Trước là AM2:AX2
    ws["AN2"] = "Gộp 4"

    # Chỉnh độ rộng cột
    ws.column_dimensions["A"].width = 5  # Cột A (STT): 5 units
    ws.column_dimensions["B"].width = 20  # Cột B (Tài xế): 20 units
    ws.column_dimensions["C"].width = 15  # Cột C (Số điện thoại): 15 units
    for col in range(4, 52):  # Từ cột D đến AY (dịch sang phải 1 cột)
        if col <= 26:
            col_letter = chr(64 + col)
        else:
            col_letter = "A" + chr(64 + col - 26)
        ws.column_dimensions[col_letter].width = 2  # 2 inches

    # Định nghĩa màu
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )  # Màu vàng
    orange_fill = PatternFill(
        start_color="FFA500", end_color="FFA500", fill_type="solid"
    )  # Màu cam

    # Tô màu từ hàng 3 trở đi (dịch sang phải 1 cột)
    for row in range(3, 101):  # Từ hàng 3 đến hàng 100
        ws[f"D{row}"].fill = yellow_fill  # Cột D (trước là C)
        ws[f"P{row}"].fill = yellow_fill  # Cột P (trước là O)
        ws[f"AB{row}"].fill = yellow_fill  # Cột AB (trước là AA)
        ws[f"AN{row}"].fill = yellow_fill  # Cột AN (trước là AM)

        ws[f"O{row}"].fill = orange_fill  # Cột O (trước là N)
        ws[f"AA{row}"].fill = orange_fill  # Cột AA (trước là Z)
        ws[f"AM{row}"].fill = orange_fill  # Cột AM (trước là AL)
        ws[f"AY{row}"].fill = orange_fill  # Cột AY (trước là AX)

    # Lưu file
    wb.save(file_path)
    if is_testing:
        print(
            f"Đã chạy ở chế độ testing: xóa và tạo lại sheet '{sheet_name}' trong {file_path}"
        )
    else:
        print(f"Đã khởi tạo/cập nhật sheet '{sheet_name}' trong {file_path}")


def initialize_driver_list(
    filename="Lenh_Dieu_Xe.xlsx", sheet_name="Tai_Xe", is_testing=False
):
    # Danh sách cột theo thứ tự
    columns = [
        "stt",
        "name",
        "cccd",
        "vehicle_id",
        "phone_number",
        "vehicle_load",
        "available",
    ]

    # Thiết lập độ rộng cho từng cột (đơn vị inches)
    width_settings = {
        "stt": 3.5,
        "name": 20,
        "cccd": 20,
        "vehicle_id": 15,
        "phone_number": 15,
        "vehicle_load": 12,
        "available": 8,
    }

    # Kiểm tra file, nếu không tồn tại thì tạo mới workbook
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = Workbook()
        # Xóa sheet mặc định nếu có
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # Nếu is_testing là True, xóa sheet Tai_Xe nếu tồn tại
    if is_testing and sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # Kiểm tra sheet Tai_Xe, nếu chưa có thì tạo mới
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]

    # Gộp ô A1:G1
    ws.merge_cells("A1:G1")
    ws["A1"] = "Thông tin tài xế"  # Có thể thay đổi tiêu đề theo ý muốn

    # Đặt header ở dòng 2
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=2, column=col_idx).value = col_name

    # # Thêm checkbox vào cột "available" ở dòng 3 (dòng đầu tiên của dữ liệu)
    # if ws.max_row < 3:
    #     ws.append([""] * (len(columns) - 1) + ["☐"])
    # else:
    #     # Đảm bảo cột available ở dòng 3 có checkbox nếu chưa có
    #     if ws.cell(row=3, column=len(columns)).value is None:
    #         ws.cell(row=3, column=len(columns)).value = "☐"

    # Đặt độ rộng cho từng cột
    col_letters = ["A", "B", "C", "D", "E", "F", "G"]
    for letter, field in zip(col_letters, width_settings):
        ws.column_dimensions[letter].width = width_settings[field]

    # Lưu file
    wb.save(filename)
    if is_testing:
        print(
            f"Đã chạy ở chế độ testing: xóa và tạo lại sheet '{sheet_name}' trong file '{filename}'."
        )
    else:
        print(
            f"Đã tạo/cập nhật sheet '{sheet_name}' trong file '{filename}' với header ở dòng 2 và checkbox ở dòng 3, cột 'available'."
        )


def sample_drivers(filename="Lenh_Dieu_Xe.xlsx", sheet_name="Tai_Xe"):
    # Danh sách tên mẫu
    sample_names = [
        "Nguyễn Văn A",
        "Trần Thị B",
        "Lê Văn C",
        "Phạm Thị D",
        "Hoàng Văn E",
    ]

    # Kiểm tra file tồn tại
    if not os.path.exists(filename):
        print(
            f"File '{filename}' không tồn tại. Vui lòng chạy initialize_driver_schedule trước."
        )
        return

    # Mở workbook
    wb = openpyxl.load_workbook(filename)

    # Kiểm tra sheet tồn tại
    if sheet_name not in wb.sheetnames:
        print(
            f"Sheet '{sheet_name}' không tồn tại. Vui lòng chạy initialize_driver_schedule trước."
        )
        return

    ws = wb[sheet_name]

    # Xác định dòng bắt đầu (sau dòng 2 là header)
    start_row = 3

    # Nếu đã có dữ liệu, tìm dòng trống tiếp theo
    if ws.max_row >= start_row:
        start_row = ws.max_row + 1

    # Tạo 5 tài xế mẫu
    for i in range(5):
        row = start_row + i
        # STT
        ws[f"A{row}"] = i + 1
        # Tên
        ws[f"B{row}"] = sample_names[i]
        # CCCD (12 chữ số ngẫu nhiên)
        ws[f"C{row}"] = "".join(random.choices(string.digits, k=12))
        # Vehicle ID (biển số xe ngẫu nhiên, ví dụ: 29A-12345)
        ws[f"D{row}"] = f"29A-{random.randint(10000, 99999)}"
        # Số điện thoại (10 chữ số, bắt đầu bằng 0)
        ws[f"E{row}"] = f"0{random.randint(100000000, 999999999)}"
        # Tải trọng xe (ngẫu nhiên từ 1.5 đến 10 tấn)
        ws[f"F{row}"] = round(random.uniform(1.5, 10), 1)
        # Available (checkbox ngẫu nhiên: ☐ hoặc ☑)
        ws[f"G{row}"] = random.choice(["☐", "☑"])

    # Lưu file
    wb.save(filename)
    print(
        f"Đã thêm 5 tài xế mẫu vào sheet '{sheet_name}' trong file '{filename}' bắt đầu từ dòng {start_row}."
    )


def copy_driver_data_to_timetable(
    file_path="Lenh_Dieu_Xe.xlsx",
    source_sheet="Tai_Xe",
    target_sheet="Driver_Timetable",
):
    # Kiểm tra file tồn tại
    if not os.path.exists(file_path):
        print(f"File '{file_path}' không tồn tại. Vui lòng tạo file trước.")
        return

    # Mở workbook
    wb = openpyxl.load_workbook(file_path)

    # Kiểm tra sự tồn tại của cả hai sheet
    if source_sheet not in wb.sheetnames:
        print(f"Sheet '{source_sheet}' không tồn tại trong file '{file_path}'.")
        return
    if target_sheet not in wb.sheetnames:
        print(f"Sheet '{target_sheet}' không tồn tại trong file '{file_path}'.")
        return

    # Lấy hai sheet
    ws_source = wb[source_sheet]  # Sheet Tai_Xe
    ws_target = wb[target_sheet]  # Sheet Driver_Timetable

    # Xác định các cột nguồn và đích
    source_columns = {"stt": "A", "name": "B", "phone_number": "E"}  # Từ Tai_Xe
    target_columns = {
        "STT": "A",
        "Tài xế": "B",
        "Số điện thoại": "C",
    }  # Sang Driver_Timetable

    # Bắt đầu từ dòng 3 (dòng dữ liệu đầu tiên)
    start_row = 3

    # Sao chép dữ liệu
    for row in range(start_row, ws_source.max_row + 1):
        # Lấy giá trị từ Tai_Xe
        stt = ws_source[f"{source_columns['stt']}{row}"].value
        name = ws_source[f"{source_columns['name']}{row}"].value
        phone = ws_source[f"{source_columns['phone_number']}{row}"].value

        # Ghi vào Driver_Timetable
        ws_target[f"{target_columns['STT']}{row}"].value = stt
        ws_target[f"{target_columns['Tài xế']}{row}"].value = name
        ws_target[f"{target_columns['Số điện thoại']}{row}"].value = phone

    # Lưu file
    wb.save(file_path)
    print(
        f"Đã sao chép dữ liệu STT, Tài xế, Số điện thoại từ sheet '{source_sheet}' sang sheet '{target_sheet}' trong file '{file_path}'."
    )


def check_driver_availability(
    file_path="Lenh_Dieu_Xe.xlsx", sheet_name="Driver_Timetable"
):
    # Kiểm tra file tồn tại
    if not os.path.exists(file_path):
        print(f"File '{file_path}' không tồn tại.")
        return None

    # Mở workbook
    wb = openpyxl.load_workbook(file_path)

    # Kiểm tra sheet tồn tại
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' không tồn tại trong file '{file_path}'.")
        return None

    ws = wb[sheet_name]

    # Danh sách cột từ D đến AY (48 cột, mỗi cột là 30 phút)
    col_letters = []
    for col in range(4, 52):  # Từ cột D (4) đến AY (51)
        if col <= 26:
            col_letter = chr(64 + col)
        else:
            col_letter = "A" + chr(64 + col - 26)
        col_letters.append(col_letter)

    # Màu rảnh: không tô màu (None), vàng (FFFFFF00), cam (FFFFA500)
    free_colors = [None, "FFFFFF00", "FFFFA500"]

    # Danh sách kết quả
    driver_availability = {}

    # Duyệt qua các dòng từ 3 trở đi (dòng dữ liệu tài xế)
    for row in range(3, ws.max_row + 1):
        driver_name = ws[f"B{row}"].value  # Cột B: Tên tài xế
        if not driver_name:  # Nếu không có tên tài xế, bỏ qua dòng này
            continue

        # Danh sách trạng thái rảnh cho từng khung giờ
        free_slots = []
        for col_idx, col_letter in enumerate(col_letters):
            cell = ws[f"{col_letter}{row}"]
            fill_color = (
                cell.fill.start_color.index
                if cell.fill.patternType == "solid"
                else None
            )
            is_free = fill_color in free_colors
            free_slots.append(is_free)

        # Tính các khoảng thời gian rảnh
        availability = []
        start_time = None
        for i in range(len(free_slots)):
            if free_slots[i] and start_time is None:  # Bắt đầu khoảng rảnh
                start_time = i * 0.5  # Chuyển sang giờ (0.5 = 30 phút)
            elif not free_slots[i] and start_time is not None:  # Kết thúc khoảng rảnh
                end_time = i * 0.5
                availability.append((start_time, end_time))
                start_time = None
            # Trường hợp cuối ngày
            if i == len(free_slots) - 1 and start_time is not None:
                availability.append((start_time, 24.0))

        # Thêm vào kết quả
        driver_availability[driver_name] = availability

    # Đóng workbook
    wb.close()

    # In kết quả
    print(f"Thời gian rảnh của các tài xế trong sheet '{sheet_name}':")
    for driver, times in driver_availability.items():
        print(f"- {driver}: {times}")

    return driver_availability


# Gọi hàm để thực thi
if __name__ == "__main__":
    # Chạy bình thường
    # initialize_driver_schedule()
    # Chạy ở chế độ testing (xóa sheet Tai_Xe nếu có và tạo lại)
    # initialize_driver_schedule(is_testing=True)
    # initialize_driver_timetable(is_testing=True)
    # sample_drivers()  # Thêm 5 tài xế mẫu
    # copy_driver_data_to_timetable()
    check_driver_availability()
