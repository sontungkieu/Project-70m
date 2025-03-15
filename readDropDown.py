from openpyxl import load_workbook


def read_dropdown_info(excel_file, sheet1_name="Sheet1", sheet2_name="Dia_Chi"):
    # Load file Excel
    wb = load_workbook(
        excel_file, data_only=True
    )  # data_only để lấy giá trị thực tế, không công thức

    # Truy cập Sheet1
    if sheet1_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet1_name}' không tồn tại trong file {excel_file}")
    ws1 = wb[sheet1_name]

    # Truy cập Sheet2
    if sheet2_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet2_name}' không tồn tại trong file {excel_file}")
    ws2 = wb[sheet2_name]

    # Lấy danh sách giá trị từ Sheet2 (cột A)
    dropdown_values = []
    address_map = {}
    for row, cell in enumerate(ws2["A"], start=1):
        if cell.value is not None:
            dropdown_values.append(cell.value)
            address_map[cell.value] = ws2[f"B{row}"].value  # Lưu ánh xạ Name -> Address

    # Kiểm tra drop-down list trong Sheet1
    results = []
    if ws1.data_validations is not None:
        for dv in ws1.data_validations.dataValidation:
            if dv.type == "list":  # Chỉ xử lý validation loại list
                # Lấy vùng ô áp dụng drop-down
                for cell_range in dv.ranges:
                    for row in ws1[cell_range]:
                        for cell in row:
                            current_value = cell.value
                            if current_value in dropdown_values:
                                index = dropdown_values.index(current_value)
                                address = address_map.get(
                                    current_value, "Không có địa chỉ"
                                )
                            else:
                                index = -1  # Giá trị không nằm trong drop-down
                                address = "N/A"

                            result = {
                                "cell": cell.coordinate,
                                "value": current_value,
                                "dropdown_list": dropdown_values,
                                "index_in_dropdown": index,
                                "address": address,
                            }
                            results.append(result)

    # In kết quả hoặc trả về
    for res in results:
        print(f"Ô: {res['cell']}")
        print(f"Giá trị hiện tại: {res['value']}")
        print(f"Danh sách drop-down: {res['dropdown_list']}")
        print(f"Số thứ tự trong drop-down: {res['index_in_dropdown']}")
        print(f"Địa chỉ: {res['address']}")
        print("-" * 50)

    return results


# Sử dụng hàm
excel_file = "kcn_dropdown.xlsx"
read_dropdown_info(excel_file)
