import ast
import json
import pandas as pd
from utilities.split_data import read_mapping

import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def read_output(filename):
    mapped_requests, mapping, inverse_mapping,node_id_to_request = read_mapping()
    df = pd.read_csv("data/destinations.csv")
    try:
        with open(filename, "r", encoding="utf-8") as file:
            output = file.read()
    except FileNotFoundError:
        print(f"Lỗi: Không tìm thấy file {filename}!")
        return None
    output = output.split("\n---")
    print(len(output))  # tách các ngày
    days = []
    json_data = {}
    for day_info in output[1:]:
        print("*"*20)
        day = day_info[:100].split()[1]
        print("*"*20)
        # day = day[2]# tách các xe
        day_info = day_info.split("\n")[11:]
        day_info = [s for s in day_info if s.strip()]  # loại bỏ các dòng trống
        # for u in day_info:
        #     print(u)
        json_day = []
        for i in range(0,len(day_info),3):
            vehicle_id = day_info[i] if "Route for vehicle" in day_info[i] else None
            string_routes = day_info[i+1] if "Node 0 (Arrival Time:" in day_info[i+1] else None
            string_max_distance = day_info[i+2] if i+2<len(day_info) and "Distance of the route:" in day_info[i+2]  else None
            string_total_distance = day_info[i] if "Total" in day_info[i] else None
            string_cumulative_historical_km = day_info[i+1] if "[" in day_info[i+1] and "]" in day_info[i+1] else None

            # print(day_info[i],i, len(day_info))
            # print(f"string_max_distance: {string_max_distance}")
            vehicle_id = int(vehicle_id.split(":")[0].split()[-1]) if vehicle_id else None
            max_distance = int(float(string_max_distance.split()[-1])) if string_max_distance else None 
            string_total_distance = int(float(string_total_distance.split()[-1])) if string_total_distance else None
            string_routes = [s.strip() for s in string_routes.split("->")] if string_routes else None
            cumulative_historical_km = ast.literal_eval(string_cumulative_historical_km.split(":")[-1]) if string_cumulative_historical_km else None
            def parse_node(s:str = 'Node 0 (Arrival Time: 0, Capacity: 0, Delivered: 0)'):
                s = s.split()
                node = s[1]
                arrival_time = int(s[4][:-1])
                capacity = int(s[6][:-1])
                delivered = int(s[8][:-1])
                request = node_id_to_request.get(str(node),None)
                # if not request:
                #     print("post_process.py:read_output:parse_node:node:",node, str(node))
                #     exit()
                node = inverse_mapping.get(node,"-1") if node!="0" else "0"
                return {
                    "node": node,
                    "destination":df["Name"].iloc[int(node)],
                    "arrival_time": arrival_time,
                    "capacity": capacity, #debug only
                    "delivered": delivered, #debug only
                    "request": request,
                }
            routes = [parse_node(s) for s in string_routes]if string_routes else None
            json_day.append( {
                "vehicle_id": vehicle_id,
                "max_distance": max_distance,
                "total_distance": string_total_distance,
                "routes": routes,
                "cumulative_historical_km": cumulative_historical_km,
            })

        json_data[day] = json_day
    return json_data

def read_and_save_json_output(
    filename=r"data\stdout_output_2025-02-19_00-00-00.txt",
):
    output = read_output(filename=filename)
    if output is not None:
        import json
        import os

        # Ensure the 'data/test' folder exists
        test_folder = os.path.join("data", "test")
        if not os.path.exists(test_folder):
            os.makedirs(test_folder)
        with open(
            os.path.join(test_folder, "2025-02-19_00-00-00.json"),
            "w",
            encoding="utf-8",
        ) as jsonfile:
            json.dump(output, jsonfile, indent=4)
    return output

# def recalculate_accumulated


def format_excel_file(file_path):
    """
    Mở file Excel bằng openpyxl và áp dụng định dạng:
      - Header: in đậm, căn giữa, nền màu, border.
      - Các ô dữ liệu: căn giữa, border, wrap_text.
      - Cài đặt chiều rộng cột (áp dụng động cho tất cả cột).
    """
    print(f"[DEBUG] Bắt đầu format file: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active

    # Định nghĩa format cho header
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="D7E4BC")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Format header (dòng 1)
    max_col = ws.max_column
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Format các ô dữ liệu (từ dòng 2 trở đi)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    # Cài đặt chiều rộng cho các cột (ở đây đặt cố định 25)
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25

    wb.save(file_path)
    print(f"[DEBUG] Đã định dạng file Excel: {file_path}")

def read_json_output_file(filename, output_dir="data/output_excel"):
    """
    Đọc file JSON đầu ra (theo cấu trúc mới), đồng thời dùng file drivers.json để map giữa id tài xế với thông tin:
      - Driver Name: trường "name"
      - Phone: trường "phone_number"
      - License Plate: trường "vehicle_id"
    
    File destinations.csv vẫn được dùng để tra cứu tên điểm đến theo node.
    
    Mỗi ngày trong JSON sẽ được xuất ra một file Excel với các cột:
      - Driver Name, Phone, License Plate, Total Distance
      - Các cột Stop 1, Stop 2, ... chứa thông tin chi tiết của từng tuyến (route) theo định dạng multiline.
      
    Tham số:
        filename: đường dẫn file JSON đầu vào.
        output_dir: đường dẫn thư mục để lưu file Excel đầu ra. Nếu không truyền vào, sẽ dùng thư mục mặc định.
    """
    script_dir = os.path.dirname(_file_)
    print(f"[DEBUG] Bắt đầu đọc file JSON: {filename}")

    with open(filename, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Đọc driver mapping từ file drivers.json với đường dẫn tương đối data/drivers.json
    mapping_path = os.path.join(script_dir, 'data', 'drivers.json')
    print(f"[DEBUG] Đọc driver mapping từ: {mapping_path}")
    with open(mapping_path, 'r', encoding='utf-8') as f:
        drivers_list = json.load(f)
    # Giả sử thứ tự của drivers_list ứng với thứ tự của các xe trong JSON đầu ra
    driver_mapping = { idx: driver for idx, driver in enumerate(drivers_list) }
    
    # Đọc file destinations (nếu cần dùng để mapping tên điểm đến theo node)
    destinations_path = os.path.join(script_dir, 'data', 'destinations.csv')
    print(f"[DEBUG] Đọc destinations từ: {destinations_path}")
    destinations_df = pd.read_csv(destinations_path)
    
    # Nếu output_dir không được cung cấp, sử dụng đường dẫn mặc định
    if output_dir is None:
        output_dir = os.path.join(script_dir, 'data', 'output', 'output_excel')
    os.makedirs(output_dir, exist_ok=True)
    print(f"[DEBUG] Thư mục output_excel: {output_dir}")
    
    # JSON mới: key là ngày, value là danh sách các xe
    for date, vehicles in data.items():
        print(f"\n[DEBUG] Xử lý ngày: {date}, số xe: {len(vehicles)}")
        vehicles_info = []
        max_stops = 0

        for vehicle in vehicles:
            vehicle_id = vehicle.get("vehicle_id")
            # Lấy tổng khoảng cách: dùng total_distance nếu có, nếu không dùng max_distance
            total_distance = vehicle.get("total_distance")
            max_distance = vehicle.get("max_distance", 0)
            distance = total_distance if total_distance is not None else max_distance
            
            # Lấy thông tin driver từ driver_mapping (giả sử vehicle_id của JSON đầu ra là chỉ số, ứng với thứ tự trong driver_mapping)
            driver_info = driver_mapping.get(vehicle_id)
            if driver_info:
                driver_name = driver_info.get("name", "NoName")
                phone = driver_info.get("phone_number", "NoPhone")
                license_plate = driver_info.get("vehicle_id", "NoPlate")
            else:
                driver_name = "NoName"
                phone = "NoPhone"
                license_plate = "NoPlate"

            stops = []
                        # Lấy danh sách routes, đảm bảo không bị None
            routes = vehicle.get("routes") or []
            for idx, route in enumerate(routes):
                node = route.get("node")
                # Lấy giá trị destination có sẵn; nếu cần, tra cứu thêm từ destinations.csv
                destination = route.get("destination")
                arrival_time = route.get("arrival_time")
                capacity = route.get("capacity", 0)
                delivered = route.get("delivered", 0)
                
                # Nếu có thông tin request, mở rộng hiển thị chi tiết
                request_info = route.get("request")
                if request_info:
                    # Lấy request_id để hiển thị
                    request_id = request_info.get('request_id', '')
                    request_details = (
                        f"Request ID: {request_info.get('request_id', '')}\n"
                        f"Name: {request_info.get('name', '')}\n"
                        f"Start: {request_info.get('start_place', '')} End: {request_info.get('end_place', '')}\n"
                        f"Weight: {request_info.get('weight', '')}\n"
                        f"Timeframe: {request_info.get('timeframe', '')}\n"
                        f"Note: {request_info.get('note', '')}\n"
                        f"Delivery Time: {request_info.get('delivery_time', '')}\n"
                        f"Status: {request_info.get('delivery_status', '')}"
                    )
                else:
                    request_details = "N/A"
                    request_id = ""
                
                # Tra cứu tên điểm đến từ destinations.csv dựa trên node (nếu có thể ép kiểu sang int)
                try:
                    node_int = int(node) + 1
                except:
                    node_int = node
                dest_info = destinations_df[destinations_df['ID'] == node_int]
                destination_name = dest_info['Name'].iloc[0] if not dest_info.empty else destination
                
                multiline_text = (
                    f"<<ID:{node}>> - {destination_name} - {arrival_time}\n"             
                    "__________________________________________________\n"
                    f"{delivered}/{capacity}\n"
                    "__________________________________________________\n"
                    f"<<request_id:{request_id}>> f{request_id}"
                    "__________________________________________________\n"
                )
                stops.append(multiline_text)

            
            if len(stops) > max_stops:
                max_stops = len(stops)
            
            vehicles_info.append({
                "driver_name": driver_name,
                "phone": phone,
                "license_plate": license_plate,
                "distance": distance,
                "stops": stops
            })
        
        # Xây dựng header cho DataFrame
        columns = ["Driver Name", "Phone", "License Plate", "Total Distance"]
        for i in range(max_stops):
            columns.append(f"Stop {i+1}")
            columns.append(" ")
        print(f"[DEBUG] columns = {columns}")

        df_rows = []
        for vinfo in vehicles_info:
            row_data = [vinfo["driver_name"], vinfo["phone"], vinfo["license_plate"], vinfo["distance"]]
            for stop_text in vinfo["stops"]:
                row_data.append(stop_text)
                row_data.append("")
            missing_stops = max_stops - len(vinfo["stops"])
            for _ in range(missing_stops):
                row_data.append("")
                row_data.append("")
            df_rows.append(row_data)
        
        df_final = pd.DataFrame(df_rows, columns=columns)
        print(f"[DEBUG] DataFrame cuối cho ngày {date}:\n{df_final.head()}")

        output_filename = os.path.join(output_dir, f"output_{date.replace('.', '-')}.xlsx")
        if os.path.exists(output_filename):
            os.remove(output_filename)
        df_final.to_excel(output_filename, index=False)
        print(f"[DEBUG] Đã xuất file Excel cho ngày {date} tại: {output_filename}")
        
        # Format file Excel
        format_excel_file(output_filename)


if __name__ == "__main__":
    # filename = r"data\stdout_output_2025-02-19_00-00-00.txt"
    # output = read_output(filename=filename)
    # print(output)
    # read_output(filename="data/output/2025-03-29_22-01-18.txt")
    read_output(filename = r"data\output\2025-03-30_00-52-50.txt")
    read_and_save_json_output(filename=r"data\output\2025-03-30_00-52-50.txt")