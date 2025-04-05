import csv,json
import os
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import *
from objects.request import Request


def read_dropdown_info(
    excel_file,
    target_cell,
    dropdown_range,
    sheet1_name="Sheet1",
    sheet2_name="Sheet2",
):
    wb = load_workbook(excel_file, data_only=True)
    if sheet1_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet1_name}' không tồn tại trong file {excel_file}")
    ws1 = wb[sheet1_name]
    if "!" in dropdown_range:
        sheet2_name, range_str = dropdown_range.split("!")
    else:
        range_str = dropdown_range
    if sheet2_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet2_name}' không tồn tại trong file {excel_file}")
    ws2 = wb[sheet2_name]

    dropdown_values = []
    address_map = {}
    for cell in ws2[range_str]:
        for c in cell:
            if c.value is not None:
                dropdown_values.append(c.value)
                col_num = c.column + 1
                col_letter = get_column_letter(col_num)
                address_cell = ws2[f"{col_letter}{c.row}"]
                address_map[c.value] = address_cell.value or "Không có địa chỉ"

    target = ws1[target_cell]
    current_value = target.value.strip() if isinstance(target.value, str) else target.value
    if current_value in dropdown_values:
        index = dropdown_values.index(current_value)
        address = address_map.get(current_value, "Không có địa chỉ")
    else:
        index = -1
        address = "N/A"

    return {
        "cell": target.coordinate,
        "value": current_value,
        "dropdown_list": dropdown_values,
        "index_in_dropdown": index,
        "address": address,
    }


def read_excel_file(file_path=os.path.join("Lenh_Dieu_Xe.xlsx"), sheet_name=TODAY):
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' không tồn tại trong file {file_path}")
    ws = wb[sheet_name]

    headers = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in ws["A4":"K4"][0]
    ]
    data = []
    row_index = 5
    while True:
        row = [cell.value for cell in ws[f"A{row_index}":f"K{row_index}"][0]]
        if pd.isna(row[1]) or row[1] is None:
            break
        data.append(row)
        row_index += 1

    df = pd.DataFrame(data, columns=headers)

    DROP_DOWN_EXT = "_DROP_DOWN_ID"
    dropdown_columns = {
        "KHÁCH HÀNG": DROP_DOWN_RANGE_DIA_CHI,
        "LOẠI XE": "CONFIG!F1:F6",
        "THỜI GIAN GIAO HÀNG": "CONFIG!AA1:AA5",
        "NV KẾ HOẠCH": "CONFIG!B3:B100",
        "THU TIỀN LUÔN": "CONFIG!AB1:AB2",
        "XUÁT HÓA ĐƠN": "CONFIG!AB1:AB2",
        "ĐÃ GIAO": "CONFIG!AB1:AB2",
        "NƠI BỐC HÀNG": "Dia_Chi!A1:A6",
    }

    for column_name in dropdown_columns.keys():
        if column_name in df.columns:
            df[column_name + DROP_DOWN_EXT] = None

    if "STT" in df.columns:
        df["STT"] = df["STT"].apply(
            lambda x: x.split()[1]
            if isinstance(x, str) and len(x.split()) > 1
            else (x.split()[0] if isinstance(x, str) and len(x.split()) > 0 else x)
        )

    errors = []
    for index, row in df.iterrows():
        missing_cols = [
            col for col in df.columns if col != df.columns[1] and pd.isna(row[col])
        ]
        if missing_cols:
            errors.append(
                f"Hàng {index + 5} thiếu dữ liệu ở cột: {', '.join(missing_cols)}"
            )

    if errors:
        print("\n⚠️ Các trường trống được phát hiện:")
        for error in errors:
            print(error)
    else:
        print("\n✅ Không có trường nào trống ngoài cột B dùng để dừng.")

    column_to_position = {name: chr(65 + i) for i, name in enumerate(headers)}
    for index, row in df.iterrows():
        for column_name, dropdown_range in dropdown_columns.items():
            if column_name in df.columns and row[column_name]:
                result = read_dropdown_info(
                    excel_file=file_path,
                    target_cell=f"{column_to_position[column_name]}{index + 5}",
                    dropdown_range=dropdown_range,
                    sheet1_name=sheet_name,
                )
                df.at[index, column_name + DROP_DOWN_EXT] = result["index_in_dropdown"]
    return df


def convert_to_object_request(df: pd.DataFrame, day: str) -> List[Request]:
    DROP_DOWN_EXT = "_DROP_DOWN_ID"
    dropdown_columns = [
        "KHÁCH HÀNG",
        "LOẠI XE",
        "THỜI GIAN GIAO HÀNG",
        "NƠI BỐC HÀNG",
        "NV KẾ HOẠCH",
        "THU TIỀN LUÔN",
        "XUÁT HÓA ĐƠN",
        "ĐÃ GIAO",
    ]

    # Danh sách khung giờ từ full_range_time
    full_range_time = [
        "S(08:00->12:00)",
        "C(13:30->17:30)",
        "T(19:00->23:00)",
        "D(00:30->04:30)",
        "Báo sau",
    ]
    # Ánh xạ chỉ số thành giờ bắt đầu và giờ kết thúc
    time_mapping = {
        0: [8, 12],  # S(08:00->12:00)
        1: [13, 17],  # C(13:30->17:30)
        2: [19, 23],  # T(19:00->23:00)
        3: [0, 4],  # D(00:30->04:30)
        4: [0, 0],  # Báo sau (giả định không có khung giờ cụ thể)
    }

    print("\nDữ liệu trước khi gán và xóa:")
    print(df.head())

    for column_name in dropdown_columns:
        if column_name in df.columns and column_name + DROP_DOWN_EXT in df.columns:
            df[column_name] = df[column_name + DROP_DOWN_EXT]

    drop_columns = [col for col in df.columns if DROP_DOWN_EXT in col]
    df = df.drop(columns=drop_columns)

    print("\nDữ liệu sau khi gán và xóa các cột _DROP_DOWN_ID:")
    print(df.head())
    print("\nDanh sách cột sau khi xử lý:")
    print(df.columns.tolist())

    requests = []
    for index, row in df.iterrows():
        try:
            print(f"\nHàng {index + 5}:")
            print(row)

            name = str(row["KHÁCH HÀNG"]) if pd.notna(row["KHÁCH HÀNG"]) else ""
            start_place = (
                [int(row["NƠI BỐC HÀNG"])]
                if pd.notna(row["NƠI BỐC HÀNG"]) and str(row["NƠI BỐC HÀNG"]).isdigit()
                else [0]
            )
            end_place = (
                [int(row["KHÁCH HÀNG"])]
                if pd.notna(row["KHÁCH HÀNG"]) and str(row["KHÁCH HÀNG"]).isdigit()
                else [0]
            )
            weight = (
                int(row["THỂ TÍCH (M3)"])
                if pd.notna(row["THỂ TÍCH (M3)"])
                and str(row["THỂ TÍCH (M3)"]).replace(".", "").isdigit()
                else 0
            )
            date = day
            # Ánh xạ timeframe từ full_range_time
            timeframe_idx = (
                int(row["THỜI GIAN GIAO HÀNG"])
                if pd.notna(row["THỜI GIAN GIAO HÀNG"])
                and str(row["THỜI GIAN GIAO HÀNG"]).isdigit()
                else 4
            )  # Mặc định là "Báo sau"
            timeframe = time_mapping.get(
                timeframe_idx, [0, 0]
            )  # Lấy khung giờ từ mapping, mặc định [0, 0] nếu không hợp lệ
            note = str(row["GHI CHÚ"]) if pd.notna(row["GHI CHÚ"]) else "."
            staff_id = (
                int(row["NV KẾ HOẠCH"])
                if pd.notna(row["NV KẾ HOẠCH"]) and str(row["NV KẾ HOẠCH"]).isdigit()
                else 0
            )
            # split_id = bool(row["STT"]) if pd.notna(row["STT"]) else True
            split_id = 0
            delivery_status = (
                int(row["ĐÃ GIAO"])
                if pd.notna(row["ĐÃ GIAO"]) and str(row["ĐÃ GIAO"]).isdigit()
                else 0
            )

            request = Request(
                name=name,
                start_place=start_place,
                end_place=end_place,
                weight=weight,
                date=date,
                timeframe=timeframe,
                note=note,
                staff_id=staff_id,
                split_id=split_id,
            )
            request.delivery_status = delivery_status
            requests.append(request)
        except Exception as e:
            print(f"Lỗi khi xử lý hàng {index + 5}: {e}")

    return requests


def excel_to_requests(file_path=os.path.join("Lenh_Dieu_Xe.xlsx"), sheet_name=TODAY):
    df = read_excel_file(file_path=file_path, sheet_name=sheet_name)
    return convert_to_object_request(df=df, day=sheet_name)

def excel_to_requests_and_save(file_path=os.path.join("Lenh_Dieu_Xe.xlsx"), sheet_name=datetime.now().strftime("%d%m%Y")):
    """
    Đọc file Excel, chuyển thành danh sách các Request objects và lưu vào file JSON.
    Args:
        file_path (str): Đường dẫn đến file Excel
        sheet_name (str): Tên sheet cần đọc (mặc định là ngày hiện tại dạng ddmmyyyy)
    Returns:
        list: Danh sách các Request objects
    """
    # Đọc file Excel và chuyển thành danh sách Request objects
    df = read_excel_file(file_path=file_path, sheet_name=sheet_name)
    requests = convert_to_object_request(df=df, day=sheet_name)
    
    # Chuẩn bị đường dẫn file JSON
    output_dir = "data/intermediate"
    os.makedirs(output_dir, exist_ok=True)  # Tạo thư mục nếu chưa tồn tại
    output_file = os.path.join(output_dir, f"{sheet_name}.json")
    
    # Chuyển các Request objects thành định dạng có thể lưu JSON
    requests_data = [request.to_dict() for request in requests]
    
    # Lưu vào file JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(requests_data, f, ensure_ascii=False, indent=4)
    
    return requests


if __name__ == "__main__":
    try:
        for i in range(len(DATES)):
            df = read_excel_file(file_path="data/input/Lenh_Dieu_Xe.xlsx",sheet_name=DATES[i])
            requests = convert_to_object_request(df, DATES[i])
            print("\nDanh sách các đối tượng Request:")
            for i, req in enumerate(requests[:]):
                print(f"Request {i + 1}:")
                print(f"  Name: {req.name}")
                print(f"  Start Place: {req.start_place}")
                print(f"  End Place: {req.end_place}")
                print(f"  Weight: {req.weight}")
                print(f"  Date: {req.date}")
                print(f"  Timeframe: {req.timeframe}")
                print(f"  Note: {req.note}")
                print(f"  Staff ID: {req.staff_id}")
                print(f"  Split ID: {req.split_id}")
                print(f"  Delivery Status: {req.delivery_status}")
                print(f"  Request ID: {req.request_id}")
    except Exception as e:
        print(f"Lỗi: {e}")
