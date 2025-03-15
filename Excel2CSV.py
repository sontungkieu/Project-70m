import pandas as pd
from openpyxl import load_workbook


def excel_sheet2_to_csv(excel_file, csv_file, sheet2_name="Dia_Chi"):
    # Load file Excel
    wb = load_workbook(excel_file, read_only=True)

    # Truy cập Sheet2
    if sheet2_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet2_name}' không tồn tại trong file {excel_file}")
    ws2 = wb[sheet2_name]

    # Lấy dữ liệu từ Sheet2
    data = []
    for row in ws2.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] is not None:  # Chỉ lấy hàng có dữ liệu ở cột A
            data.append({"Name": row[0], "Address": row[1] if row[1] else ""})

    if not data:
        raise ValueError(f"Sheet2 trong {excel_file} không có dữ liệu!")

    # Tạo DataFrame và thêm cột ID
    df = pd.DataFrame(data)
    df.insert(0, "ID", range(1, len(df) + 1))  # Thêm cột ID từ 1

    # Lưu vào file CSV
    df.to_csv(csv_file, index=False, encoding="utf-8")
    print(f"Đã chuyển dữ liệu từ {excel_file}, Sheet2 sang {csv_file}!")


# Sử dụng hàm
excel_file = "kcn_dropdown.xlsx"
csv_file = "destinations_output.csv"
excel_sheet2_to_csv(excel_file, csv_file)
